from __future__ import annotations

import json
import csv
import re
import sys
import time
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import fornies.scripts.build_supplier_audit_panel_v05 as v05


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
PHASE_01_DIR = BASE_DIR / "output" / "fase_01_cadastro"
PHASE_03_DIR = BASE_DIR / "output" / "fase_03_enriquecimento"
PHASE_04_DIR = BASE_DIR / "output" / "fase_04_visualizacao"
PHASE_06_DIR = BASE_DIR / "output" / "fase_06_curva_total"

CURVE_FILE = DATA_DIR / "CURVA_ABC_FORN_-_TOTAL.xlsx"
MASTER_FILE = DATA_DIR / "FORNECEDORES - TODAS AS EMPRESAS.xlsx"
NORMALIZED_FILE = PHASE_01_DIR / "01_fornecedores_cadastro_normalizado.csv"
CACHE_DIR = PHASE_03_DIR / "99_cache_opencnpj"
HTML_FILE = PHASE_04_DIR / "06_painel_auditoria_fornecedores_curva_total.html"

OUT_00_SUMMARY = PHASE_06_DIR / "00_resumo_versao_06.json"
OUT_01_CURVE = PHASE_06_DIR / "01_curva_total_consolidada.csv"
OUT_02_INTERNAL = PHASE_06_DIR / "02_fornecedores_curva_total_com_cadastro_endereco.csv"
OUT_03_OPENCNPJ = PHASE_06_DIR / "03_fornecedores_curva_total_opencnpj.csv"
OUT_04_AUDIT = PHASE_06_DIR / "04_fornecedores_curva_total_auditoria.csv"
OUT_05_PENDING = PHASE_06_DIR / "05_fornecedores_curva_total_pendencias.csv"

CURVE_ORDER = {
    "AAA": 1,
    "AA": 2,
    "A": 3,
    "B": 4,
    "BB": 5,
    "C": 6,
    "CC": 7,
    "CCC": 8,
}


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def clean_upper(value: Any) -> str:
    return clean(value).upper()


def digits(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def as_float(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        try:
            return float(value)
        except Exception:
            return 0.0


def strip_accents(value: Any) -> str:
    text = clean_upper(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9 ]+", " ", text).strip()


def code_type(code: str) -> str:
    prefix = clean_upper(code)[:1]
    if prefix == "J":
        return "PJ/CNPJ"
    if prefix == "F":
        return "PF/CPF"
    return "Outro"


def cnpj_from_code(code: str) -> str:
    if clean_upper(code).startswith("J"):
        doc = digits(code)
        return doc if len(doc) == 14 else ""
    return ""


def best_value(values: list[Any]) -> str:
    for value in values:
        text = clean(value)
        if text:
            return text
    return ""


def join_unique(values: list[Any], sep: str = ";") -> str:
    seen: list[str] = []
    for value in values:
        text = clean(value)
        if text and text not in seen:
            seen.append(text)
    return sep.join(seen)


def row_completeness(row: dict[str, Any], fields: list[str]) -> int:
    return sum(1 for field in fields if clean(row.get(field)))


def read_excel_all_sheets(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        iterator = ws.iter_rows(values_only=True)
        header = next(iterator, None)
        if not header:
            continue
        columns = [clean(col) for col in header]
        for excel_row_idx, values in enumerate(iterator, start=2):
            row = {columns[idx]: values[idx] if idx < len(values) else "" for idx in range(len(columns))}
            row["EMPRESA"] = sheet_name
            row["_linha_planilha"] = str(excel_row_idx)
            rows.append(row)
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def find_address_file() -> Path:
    matches = sorted(DATA_DIR.glob("FORNECEDORES ENDERE*EMPRESAS.xlsx"))
    if not matches:
        raise FileNotFoundError("Planilha de enderecos nao encontrada em data/FORNECEDORES ENDERE*EMPRESAS.xlsx")
    return matches[0]


def format_internal_address(row: dict[str, Any]) -> str:
    parts = [
        clean(row.get("DSENDEFORN")),
        clean(row.get("NRENDEFORNSEP")),
        clean(row.get("DSENDEFORNSEP")),
        clean(row.get("DSCOMPENDFOR")),
        clean(row.get("NMBAIRFORN")),
        clean(row.get("NRCEPFORN")),
    ]
    return " - ".join(part for part in parts if part)


def format_official_address(payload: dict[str, Any] | None) -> str:
    if not isinstance(payload, dict):
        return ""
    parts = [
        clean(payload.get("tipo_logradouro")),
        clean(payload.get("logradouro")),
        clean(payload.get("numero")),
        clean(payload.get("complemento")),
        clean(payload.get("bairro")),
        clean(payload.get("cep")),
    ]
    return " ".join(part for part in parts if part)


def collect_internal_phones(row: dict[str, Any]) -> str:
    fields = ["NRTELEFORN", "NRTELEFORN1", "NRCELULARFOR", "NRTELRESIFOR", "NRFAXFORN"]
    return join_unique([row.get(field) for field in fields])


def collect_official_phones(payload: dict[str, Any] | None) -> str:
    if not isinstance(payload, dict):
        return ""
    phones: list[str] = []
    for item in payload.get("telefones") or []:
        if not isinstance(item, dict):
            continue
        phone = f"{clean(item.get('ddd'))}{clean(item.get('numero'))}"
        if phone:
            phones.append(phone)
    return join_unique(phones)


def cache_file(cnpj: str) -> Path:
    return CACHE_DIR / f"{cnpj}.json"


def load_cached(cnpj: str) -> dict[str, Any] | None:
    path = cache_file(cnpj)
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def save_cached(cnpj: str, payload: dict[str, Any]) -> None:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_file(cnpj).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def build_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=2,
        backoff_factor=0.8,
        status_forcelist=[408, 429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update(
        {
            "accept": "application/json",
            "user-agent": "Fornecedores-v06-OpenCNPJ/1.0",
            "accept-language": "pt-BR,pt;q=0.9",
        }
    )
    return session


def progress_line(done: int, total: int, counts: Counter[str], current: str = "") -> None:
    width = 28
    ratio = 1 if total == 0 else done / total
    filled = min(width, int(width * ratio))
    bar = "#" * filled + "-" * (width - filled)
    text = (
        f"\rOpenCNPJ [{bar}] {done}/{total} {ratio:6.1%} "
        f"cache={counts['cache']} api={counts['api']} erro={counts['erro']} "
        f"nao_aplicavel={counts['nao_aplicavel']} {current[:18]:18}"
    )
    sys.stdout.write(text)
    sys.stdout.flush()


def fetch_opencnpj_payloads(cnpjs: list[str], delay_seconds: float = 0.4) -> dict[str, tuple[str, dict[str, Any] | None]]:
    session = build_session()
    results: dict[str, tuple[str, dict[str, Any] | None]] = {}
    counts: Counter[str] = Counter()
    total = len(cnpjs)
    progress_line(0, total, counts)

    for idx, cnpj in enumerate(cnpjs, start=1):
        cached = load_cached(cnpj)
        if cached is not None and not str(cached.get("__status_api", "")).startswith("http_"):
            status = "cache"
            payload = cached
            counts["cache"] += 1
        else:
            try:
                response = session.get(f"https://api.opencnpj.org/{cnpj}", timeout=20)
                if response.status_code == 200:
                    payload = response.json()
                    if isinstance(payload, dict):
                        status = "api"
                        save_cached(cnpj, payload)
                        counts["api"] += 1
                    else:
                        status = "payload_invalido"
                        payload = None
                        counts["erro"] += 1
                else:
                    status = f"http_{response.status_code}"
                    payload = {"__status_api": status, "cnpj": cnpj}
                    save_cached(cnpj, payload)
                    counts["erro"] += 1
            except Exception:
                status = "erro_requisicao"
                payload = None
                counts["erro"] += 1
            if delay_seconds > 0:
                time.sleep(delay_seconds)

        results[cnpj] = (status, payload)
        if idx == total or idx == 1 or idx % 25 == 0 or status not in {"cache"}:
            progress_line(idx, total, counts, cnpj)

    sys.stdout.write("\n")
    return results


def infer_regime(payload: dict[str, Any] | None) -> str:
    if not isinstance(payload, dict):
        return "indeterminado"
    if clean_upper(payload.get("opcao_mei")) == "S":
        return "mei"
    simples = clean_upper(payload.get("opcao_simples"))
    if simples == "S":
        return "simples_nacional"
    if simples == "N":
        return "regime_nao_simples"
    return "indeterminado"


def infer_credit(payload: dict[str, Any] | None) -> str:
    regime = infer_regime(payload)
    if regime == "regime_nao_simples":
        return "alto_potencial"
    if regime == "simples_nacional":
        return "potencial_condicionado"
    if regime == "mei":
        return "baixo_ou_inexistente"
    return "indeterminado"


def compare_place(internal_value: str, official_value: str) -> str:
    if clean(internal_value) and clean(official_value):
        return "OK" if strip_accents(internal_value) == strip_accents(official_value) else "Divergencia"
    if not clean(internal_value) and clean(official_value):
        return "Dado novo"
    if clean(internal_value) and not clean(official_value):
        return "Sem dado oficial"
    return "Sem dado"


def compare_uf(internal_value: str, official_value: str) -> str:
    left = clean_upper(internal_value)
    right = clean_upper(official_value)
    if left and right:
        return "OK" if left == right else "Divergencia"
    if not left and right:
        return "Dado novo"
    if left and not right:
        return "Sem dado oficial"
    return "Sem dado"


def classify_status(status: str, api_status: str) -> str:
    text = clean(status).lower()
    if text == "ativa":
        return "ativa"
    if text:
        return "nao_ativa"
    if api_status in {"api", "cache"}:
        return "sem_status_retorno"
    return "nao_consultado"


def classify_regime_status(regime: str) -> str:
    if regime == "regime_nao_simples":
        return "definido_nao_simples"
    if regime == "simples_nacional":
        return "definido_simples"
    if regime == "mei":
        return "definido_mei"
    return "indeterminado"


def read_sheet_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    header = next(iterator, None)
    if not header:
        return []
    columns = [clean(col) for col in header]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        rows.append({columns[idx]: values[idx] if idx < len(values) else "" for idx in range(len(columns))})
    return rows


def build_curve() -> tuple[list[dict[str, Any]], int]:
    curve_rows = read_sheet_rows(CURVE_FILE)
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in curve_rows:
        row = row.copy()
        row["CDFORNECED"] = clean_upper(row.get("CDFORNECED"))
        row["TOT_FORN_NUM"] = as_float(row.get("TOT_FORN"))
        groups[row["CDFORNECED"]].append(row)

    total_value = float(sum(row["TOT_FORN_NUM"] for group in groups.values() for row in group))
    records: list[dict[str, Any]] = []

    for code, group in groups.items():
        best = min(group, key=lambda row: CURVE_ORDER.get(clean_upper(row.get("CURVA")), 99))
        value = float(sum(as_float(row.get("TOT_FORN_NUM")) for row in group))
        detail_rows = [{key: value for key, value in row.items() if key != "TOT_FORN_NUM"} for row in group]
        records.append(
            {
                "codigo": code,
                "tipo": code_type(code),
                "cnpj_codigo": cnpj_from_code(code),
                "razao_curva": best_value([row.get("RAZAO_SOCIAL") for row in group]),
                "valor_curva_total": value,
                "perc_curva": (value / total_value * 100) if total_value else 0,
                "curva_original_referencia": clean_upper(best.get("CURVA")),
                "pos_original_referencia": clean(best.get("POS")),
                "linhas_curva": len(group),
                "linhas_curva_detalhe": json.dumps(detail_rows, ensure_ascii=False),
            }
        )

    records.sort(key=lambda row: row["valor_curva_total"], reverse=True)
    accumulated = 0.0
    for index, row in enumerate(records, start=1):
        accumulated += row["perc_curva"]
        row["posicao_curva_consolidada"] = index
        row["perc_acumulado_consolidado"] = accumulated
        row["curva_consolidada"] = row["curva_original_referencia"]
    return records, len(curve_rows)


def build_indexes() -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]], dict[str, list[dict[str, str]]]]:
    master_by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in read_excel_all_sheets(MASTER_FILE):
        row["CDFORNECED"] = clean_upper(row.get("CDFORNECED"))
        row["NRINSJURFORN_DOC"] = digits(row.get("NRINSJURFORN"))
        master_by_code[row["CDFORNECED"]].append(row)

    address_by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in read_excel_all_sheets(find_address_file()):
        row["CDFORNECED"] = clean_upper(row.get("CDFORNECED"))
        address_by_code[row["CDFORNECED"]].append(row)

    normalized_by_code: dict[str, list[dict[str, str]]] = defaultdict(list)
    if NORMALIZED_FILE.exists():
        with NORMALIZED_FILE.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                normalized_by_code[clean_upper(row.get("codigo_fornecedor"))].append(row)

    return master_by_code, address_by_code, normalized_by_code


def choose_best_address(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {}
    fields = ["SGESTADO", "NMMUNICIPIO", "CDMUNICIBGE", "DSENDEFORN", "NMBAIRFORN", "NRCEPFORN", "NRTELEFORN", "DSEMAILFORN", "NMCONTFORN"]
    return max(rows, key=lambda row: row_completeness(row, fields))


def choose_best_normalized(rows: list[dict[str, str]]) -> dict[str, str]:
    if not rows:
        return {}
    return max(rows, key=lambda row: as_float(row.get("score_completude")))


def build_v06_rows() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    PHASE_06_DIR.mkdir(parents=True, exist_ok=True)
    curve, original_curve_rows = build_curve()
    write_csv(OUT_01_CURVE, curve)

    master_by_code, address_by_code, normalized_by_code = build_indexes()
    valid_cnpjs = sorted({row["cnpj_codigo"] for row in curve if len(row["cnpj_codigo"]) == 14})
    opencnpj = fetch_opencnpj_payloads(valid_cnpjs)

    rows: list[dict[str, Any]] = []
    internal_records: list[dict[str, Any]] = []
    opencnpj_records: list[dict[str, Any]] = []
    pending_records: list[dict[str, Any]] = []

    for curve_row in curve:
        code = clean_upper(curve_row["codigo"])
        master_rows = master_by_code.get(code, [])
        address_rows = address_by_code.get(code, [])
        normalized_rows = normalized_by_code.get(code, [])
        normalized = choose_best_normalized(normalized_rows)
        address = choose_best_address(address_rows)

        companies = sorted({clean_upper(row.get("EMPRESA")) for row in master_rows if clean(row.get("EMPRESA"))})
        cnpj = curve_row["cnpj_codigo"] or best_value([row.get("NRINSJURFORN_DOC") for row in master_rows if len(digits(row.get("NRINSJURFORN_DOC"))) == 14])
        api_status, payload = opencnpj.get(cnpj, ("nao_aplicavel", None)) if cnpj else ("nao_aplicavel", None)
        if curve_row["tipo"] != "PJ/CNPJ":
            api_status, payload = "nao_aplicavel", None

        razao_internal = best_value([row.get("NMRAZSOCFORN") for row in master_rows]) or curve_row["razao_curva"]
        fantasia_internal = best_value([row.get("NMFANTFORN") for row in master_rows])
        status_internal = best_value([row.get("IDSITUCADA") for row in master_rows])
        ie_internal = best_value([normalized.get("inscricao_estadual"), best_value([row.get("NRINSCESFORN") for row in master_rows])])
        im_internal = best_value([normalized.get("inscricao_municipal"), best_value([row.get("NRINSJMUNF") for row in master_rows])])
        internal_email = join_unique([normalized.get("emails"), address.get("DSEMAILFORN")])
        internal_phone = collect_internal_phones(address)
        official_phone = collect_official_phones(payload)
        official_email = clean(payload.get("email")) if isinstance(payload, dict) else ""
        official_uf = clean(payload.get("uf")) if isinstance(payload, dict) else ""
        official_city = clean(payload.get("municipio")) if isinstance(payload, dict) else ""
        internal_uf = clean(address.get("SGESTADO"))
        internal_city = clean(address.get("NMMUNICIPIO"))
        uf_status = compare_uf(internal_uf, official_uf) if cnpj else "Nao aplicavel"
        city_status = compare_place(internal_city, official_city) if cnpj else "Nao aplicavel"
        official_address = format_official_address(payload)
        internal_address = format_internal_address(address)
        regime = infer_regime(payload)
        credit = infer_credit(payload)

        divergences = []
        if isinstance(payload, dict) and clean(payload.get("razao_social")) and strip_accents(razao_internal) != strip_accents(payload.get("razao_social")):
            divergences.append("razao_social")
        if isinstance(payload, dict) and clean(payload.get("nome_fantasia")) and fantasia_internal and strip_accents(fantasia_internal) != strip_accents(payload.get("nome_fantasia")):
            divergences.append("nome_fantasia")
        if official_email and internal_email and official_email.lower() not in internal_email.lower():
            divergences.append("email")
        if uf_status == "Divergencia":
            divergences.append("uf")
        if city_status == "Divergencia":
            divergences.append("municipio")

        missing = []
        if not master_rows:
            missing.append("fornecedor_fora_cadastro")
        if not address_rows:
            missing.append("endereco_interno")
        if not internal_uf:
            missing.append("uf")
        if not internal_city:
            missing.append("municipio")
        if not clean(address.get("NRCEPFORN")):
            missing.append("cep")
        if not internal_phone:
            missing.append("telefone")
        if not internal_email:
            missing.append("email")

        new_data = []
        if official_uf and not internal_uf:
            new_data.append("uf_oficial")
        if official_city and not internal_city:
            new_data.append("municipio_oficial")
        if official_address and not internal_address:
            new_data.append("endereco_oficial")
        if official_phone and not internal_phone:
            new_data.append("telefone_oficial")
        if official_email and not internal_email:
            new_data.append("email_oficial")
        if internal_address:
            new_data.append("endereco_interno")
        if internal_phone:
            new_data.append("telefone_interno")

        needs_saneamento = bool(missing)
        needs_validation = bool(divergences) or regime == "indeterminado"
        has_new_data = bool(new_data) or isinstance(payload, dict)
        any_divergence = bool(divergences)

        action = "Revisar antes" if any_divergence else "Validar fiscal" if needs_validation else "Sanear cadastro" if needs_saneamento else "Atualizar BD" if has_new_data else "Sem alteracao"
        score = best_value([normalized.get("score_completude")]) or ("0" if not master_rows else "50")
        score_float = as_float(score)
        faixa = best_value([normalized.get("faixa_completude")])
        if not faixa:
            faixa = "alta" if score_float >= 85 else "media" if score_float >= 60 else "baixa"

        row = {
            "linha_original": "",
            "empresa": companies[0] if companies else "FORA_CADASTRO",
            "codigo": code,
            "tipo": curve_row["tipo"],
            "documento": cnpj or digits(code),
            "chave": code,
            "razao_original": razao_internal,
            "fantasia_original": fantasia_internal or "Sem dado",
            "status_interno": status_internal,
            "emails_originais": internal_email,
            "ie_original": ie_internal or "Sem dado",
            "im_original": im_internal or "Sem dado",
            "razao_social_oficial": clean(payload.get("razao_social")) if isinstance(payload, dict) else "",
            "nome_fantasia_oficial": clean(payload.get("nome_fantasia")) if isinstance(payload, dict) else "",
            "situacao_cadastral_oficial": clean(payload.get("situacao_cadastral")) if isinstance(payload, dict) else "",
            "email_oficial": official_email,
            "telefones_oficiais": internal_phone or official_phone or "Sem dado",
            "endereco_completo": internal_address or official_address or "Sem dado",
            "endereco_municipio": internal_city or official_city or "Sem dado",
            "endereco_uf": internal_uf or official_uf or "Sem UF",
            "endereco_cep": clean(address.get("NRCEPFORN")) or (clean(payload.get("cep")) if isinstance(payload, dict) else ""),
            "contatos_qsa": clean(address.get("NMCONTFORN")) or "Sem dado",
            "matriz_filial": clean(payload.get("matriz_filial")) if isinstance(payload, dict) else "",
            "porte_empresa": clean(payload.get("porte_empresa")) if isinstance(payload, dict) else "",
            "natureza_juridica": clean(payload.get("natureza_juridica")) if isinstance(payload, dict) else "",
            "opcao_simples": clean(payload.get("opcao_simples")) if isinstance(payload, dict) else "",
            "opcao_mei": clean(payload.get("opcao_mei")) if isinstance(payload, dict) else "",
            "status_api": api_status,
            "regime_indiciado": regime,
            "potencial_credito_2027": credit,
            "status_cadastro_oficial": classify_status(clean(payload.get("situacao_cadastral")) if isinstance(payload, dict) else "", api_status),
            "status_regime_fiscal": classify_regime_status(regime),
            "necessita_saneamento_cadastral": "S" if needs_saneamento else "N",
            "motivos_saneamento_cadastral": ";".join(missing),
            "necessita_validacao_fiscal": "S" if needs_validation else "N",
            "motivos_validacao_fiscal": ";".join(divergences),
            "divergencia_razao_social": "S" if "razao_social" in divergences else "N",
            "divergencia_nome_fantasia": "S" if "nome_fantasia" in divergences else "N",
            "divergencia_email": "S" if "email" in divergences else "N",
            "movimento_nfe": "S",
            "nfe_linhas": str(int(curve_row["linhas_curva"])),
            "nfe_valor_total": curve_row["valor_curva_total"],
            "nfe_produtos_unicos": "",
            "nfe_ufs": internal_uf or official_uf or "Sem UF",
            "nfe_anos": "2024;2025;2026",
            "curva": curve_row["curva_consolidada"],
            "curva_posicao": str(int(curve_row["posicao_curva_consolidada"])),
            "curva_total_fornecedor": curve_row["valor_curva_total"],
            "curva_percentual": curve_row["perc_curva"],
            "score_cadastral": score,
            "faixa_completude": faixa,
            "pendencias_essenciais": normalized.get("pendencias_essenciais", ""),
            "pendencias_complementares": normalized.get("pendencias_complementares", ""),
            "prioridade_saneamento_score": "",
            "prioridade_saneamento_faixa": "",
            "origem_cadastro_label": "Cadastrado" if master_rows else "Fora do cadastro",
            "endereco_status_label": "Sem endereco" if not address_rows else "Completo" if internal_uf and internal_city and clean(address.get("NRCEPFORN")) and internal_address else "Incompleto",
            "uf_status": uf_status,
            "municipio_status": city_status,
            "endereco_oficial": official_address,
            "telefone_oficial": official_phone,
            "telefone_interno": internal_phone,
            "dados_novos_v06": ";".join(new_data),
            "divergencias_v06": ";".join(divergences),
            "linha_curva_detalhe": curve_row["linhas_curva_detalhe"],
        }

        summary = []
        if new_data:
            summary.append(f"{len(new_data)} incluidos")
        if divergences:
            summary.append(f"{len(divergences)} divergencias")
        if not summary:
            summary.append("sem dado novo")

        row.update(
            {
                "nome_exibicao": row["razao_original"] or row["razao_social_oficial"] or code,
                "status_interno_label": v05.internal_status_label(row["status_interno"]),
                "situacao_label": v05.official_status_label(row["situacao_cadastral_oficial"], row["status_api"]),
                "regime_label": v05.short_regime(row["regime_indiciado"]),
                "credito_label": v05.short_credit(row["potencial_credito_2027"]),
                "movimento_label": "Com movimento",
                "curva_label": row["curva"] or "Sem curva",
                "uf_label": row["endereco_uf"] or "Sem UF",
                "curva_valor_fmt": v05.money(row["curva_total_fornecedor"]),
                "curva_pct_fmt": v05.pct(row["curva_percentual"]),
                "nfe_valor_fmt": v05.money(row["nfe_valor_total"]),
                "dado_novo_label": "Com dado novo" if has_new_data else "Sem dado novo",
                "divergencia_label": "Qualquer divergencia" if any_divergence else "Sem divergencia",
                "saneamento_label": v05.yes_no_label(row["necessita_saneamento_cadastral"], "Precisa sanear", "Nao precisa"),
                "validacao_label": v05.yes_no_label(row["necessita_validacao_fiscal"], "Precisa validar", "Nao precisa"),
                "acao_label": action,
                "mudanca_resumo": " / ".join(summary),
                "sort_curva_rank": CURVE_ORDER.get(row["curva"], 99),
                "sort_curva_valor": as_float(row["curva_total_fornecedor"]),
                "sort_nfe_valor": as_float(row["nfe_valor_total"]),
                "row_tone": v05.action_to_tone(action, has_new_data),
                "empresas_lista": companies or ["Fora do cadastro"],
                "empresas_label": ";".join(companies) if companies else "Fora do cadastro",
                "codigos_lista": [code],
                "codigos_label": code,
                "ocorrencias": [
                    {
                        "empresa": clean(row_m.get("EMPRESA")),
                        "codigo": clean_upper(row_m.get("CDFORNECED")),
                        "linha_original": clean(row_m.get("_linha_planilha")),
                        "razao_original": clean(row_m.get("NMRAZSOCFORN")),
                        "fantasia_original": clean(row_m.get("NMFANTFORN")),
                        "status_interno": clean(row_m.get("IDSITUCADA")),
                    }
                    for row_m in master_rows
                ],
                "documento_formatado": v05.format_document(row["documento"]),
                "row_key": code,
            }
        )

        rows.append(row)
        internal_records.append(
            {
                "codigo": code,
                "origem_cadastro": row["origem_cadastro_label"],
                "empresas": row["empresas_label"],
                "cnpj": cnpj,
                "razao_interna": razao_internal,
                "endereco_status": row["endereco_status_label"],
                "uf_interna": internal_uf,
                "municipio_interno": internal_city,
                "telefone_interno": internal_phone,
                "email_interno": internal_email,
            }
        )
        opencnpj_records.append(
            {
                "codigo": code,
                "cnpj": cnpj,
                "status_api": api_status,
                "razao_oficial": row["razao_social_oficial"],
                "situacao_oficial": row["situacao_cadastral_oficial"],
                "uf_oficial": official_uf,
                "municipio_oficial": official_city,
                "uf_status": uf_status,
                "municipio_status": city_status,
                "regime": row["regime_label"],
                "credito": row["credito_label"],
            }
        )
        if missing or divergences:
            pending_records.append(
                {
                    "codigo": code,
                    "cnpj": cnpj,
                    "razao": row["nome_exibicao"],
                    "pendencias": ";".join(missing),
                    "divergencias": ";".join(divergences),
                    "acao": action,
                }
            )

    rows.sort(key=lambda item: (item["sort_curva_rank"], -item["sort_curva_valor"], item["nome_exibicao"]))
    write_csv(OUT_02_INTERNAL, internal_records)
    write_csv(OUT_03_OPENCNPJ, opencnpj_records)
    write_csv(
        OUT_04_AUDIT,
        [{key: json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value for key, value in row.items()} for row in rows],
    )
    write_csv(OUT_05_PENDING, pending_records)

    summary = {
        "versao": "06",
        "universo": "fornecedores presentes na curva ABC total",
        "arquivo_curva": str(CURVE_FILE.relative_to(BASE_DIR)),
        "fornecedores_curva_consolidados": len(rows),
        "linhas_curva_original": original_curve_rows,
        "valor_total_curva_consolidado": float(sum(row["sort_curva_valor"] for row in rows)),
        "fornecedores_cadastrados": sum(1 for row in rows if row["origem_cadastro_label"] == "Cadastrado"),
        "fornecedores_fora_cadastro": sum(1 for row in rows if row["origem_cadastro_label"] == "Fora do cadastro"),
        "fornecedores_com_endereco_completo": sum(1 for row in rows if row["endereco_status_label"] == "Completo"),
        "fornecedores_com_endereco_incompleto": sum(1 for row in rows if row["endereco_status_label"] == "Incompleto"),
        "fornecedores_sem_endereco": sum(1 for row in rows if row["endereco_status_label"] == "Sem endereco"),
        "cnpjs_validos_opencnpj": len(valid_cnpjs),
        "opencnpj_status": dict(Counter(row["status_api"] for row in rows)),
        "uf_status": dict(Counter(row["uf_status"] for row in rows)),
        "municipio_status": dict(Counter(row["municipio_status"] for row in rows)),
        "outputs": [
            str(OUT_01_CURVE.relative_to(BASE_DIR)),
            str(OUT_02_INTERNAL.relative_to(BASE_DIR)),
            str(OUT_03_OPENCNPJ.relative_to(BASE_DIR)),
            str(OUT_04_AUDIT.relative_to(BASE_DIR)),
            str(OUT_05_PENDING.relative_to(BASE_DIR)),
            str(HTML_FILE.relative_to(BASE_DIR)),
        ],
    }
    OUT_00_SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return rows, summary


def build_html(rows: list[dict[str, Any]]) -> None:
    PHASE_04_DIR.mkdir(parents=True, exist_ok=True)
    original_target = v05.AUDIT_HTML_FILE
    v05.AUDIT_HTML_FILE = HTML_FILE
    try:
        v05.build_html(rows)
    finally:
        v05.AUDIT_HTML_FILE = original_target

    html = HTML_FILE.read_text(encoding="utf-8")
    html = html.replace("<title>Painel de Auditoria de Fornecedores</title>", "<title>Painel de Fornecedores - Versao 06</title>")
    html = html.replace("<h1>Painel de Fornecedores</h1>", "<h1>Painel de Fornecedores - Curva Total</h1>")
    html = html.replace(
        "Visualizacao interativa do cadastro, movimento, enriquecimento e leitura fiscal inicial.",
        "Versao 06: fornecedores da curva ABC total, com enderecos internos, cache/API OpenCNPJ e comparacao de UF/municipio.",
    )
    HTML_FILE.write_text(html, encoding="utf-8")


def update_diary(summary: dict[str, Any]) -> None:
    diary = BASE_DIR / "docs" / "DIARIO_DE_BORDO.md"
    current = diary.read_text(encoding="utf-8") if diary.exists() else ""
    marker = "#### Implementacao da versao 06 - Curva total"
    if marker in current:
        return
    entry = f"""

{marker}

- Foi criado o script `scripts/build_supplier_audit_panel_v06.py`.
- A versao `06` parte logicamente da versao `05`, mas gera arquivos proprios e nao altera as versoes `04` e `05`.
- Foram gerados arquivos intermediarios numerados em `output/fase_06_curva_total`.
- O universo considerado foi a curva `data/CURVA_ABC_FORN_-_TOTAL.xlsx`, consolidada por `CDFORNECED`.
- Fornecedores consolidados na versao `06`: `{summary['fornecedores_curva_consolidados']}`.
- Fornecedores cadastrados no cadastro mestre: `{summary['fornecedores_cadastrados']}`.
- Fornecedores da curva fora do cadastro mestre: `{summary['fornecedores_fora_cadastro']}`.
- CNPJs validos tratados para OpenCNPJ: `{summary['cnpjs_validos_opencnpj']}`.
- A consulta OpenCNPJ usa cache local primeiro e consulta API somente quando o CNPJ valido nao esta cacheado.
- O script exibe uma barra de progresso no terminal durante a resolucao OpenCNPJ, com contadores de `cache`, `api`, `erro` e percentual concluido.
- Foi gerado o painel `output/fase_04_visualizacao/06_painel_auditoria_fornecedores_curva_total.html`.
"""
    with diary.open("a", encoding="utf-8") as handle:
        handle.write(entry)


def main() -> None:
    rows, summary = build_v06_rows()
    build_html(rows)
    update_diary(summary)
    print(f"HTML gerado: {HTML_FILE}")
    print(f"Resumo: {OUT_00_SUMMARY}")
    print(f"Registros: {len(rows)}")


if __name__ == "__main__":
    main()
