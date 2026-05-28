from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
PHASE_01_DIR = BASE_DIR / "output" / "fase_01_cadastro"
PHASE_02_DIR = BASE_DIR / "output" / "fase_02_reconciliacao"
PHASE_03_DIR = BASE_DIR / "output" / "fase_03_enriquecimento"
PHASE_04_DIR = BASE_DIR / "output" / "fase_04_visualizacao"

MASTER_FILE = DATA_DIR / "FORNECEDORES - TODAS AS EMPRESAS.xlsx"
NORMALIZED_FILE = PHASE_01_DIR / "01_fornecedores_cadastro_normalizado.csv"
RECONCILED_FILE = PHASE_02_DIR / "01_fornecedores_reconciliados_movimento.csv"
CLASSIFIED_FILE = PHASE_03_DIR / "03_fornecedores_classificacao_fiscal_inicial_lote_completo.csv"
ENRICHED_FILE = PHASE_03_DIR / "01_fornecedores_opencnpj_enriquecidos_lote_completo.csv"

AUDIT_HTML_FILE = PHASE_04_DIR / "04_painel_auditoria_fornecedores.html"

CURVE_ORDER = {
    "AAA": 1,
    "AA": 2,
    "A": 3,
    "B": 4,
    "BB": 5,
    "C": 6,
    "CC": 7,
    "CCC": 8,
}


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_text_upper(value: Any) -> str:
    return normalize_text(value).upper()


def normalize_doc(value: Any, expected_len: int | None = None) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if expected_len and len(digits) != expected_len:
        return ""
    return digits


def as_float(value: Any) -> float:
    try:
        return float(str(value or "0").replace(",", "."))
    except ValueError:
        return 0.0


def money(value: Any) -> str:
    amount = as_float(value)
    if amount <= 0:
        return "R$ 0"
    return "R$ " + f"{amount:,.0f}".replace(",", ".")


def format_document(value: Any) -> str:
    digits = normalize_doc(value)
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return digits


def pct(value: Any) -> str:
    amount = as_float(value)
    if amount <= 0:
        return ""
    return f"{amount:.2f}%".replace(".", ",")


def load_csv_by_field(path: Path, field: str) -> dict[str, dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return {row[field]: row for row in csv.DictReader(handle)}


def load_csv_index(path: Path, key_fields: list[str]) -> dict[tuple[str, ...], dict[str, str]]:
    out: dict[tuple[str, ...], dict[str, str]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            out[tuple(row[field] for field in key_fields)] = row
    return out


def classify_code_type(code: str) -> str:
    prefix = str(code or "").strip()[:1].upper()
    if prefix == "J":
        return "PJ/CNPJ"
    if prefix == "F":
        return "PF/CPF"
    return "Outro"


def short_regime(value: str) -> str:
    text = normalize_text(value).lower()
    if text == "simples_nacional":
        return "Simples"
    if text == "regime_nao_simples":
        return "Nao Simples"
    if text == "mei":
        return "MEI"
    if text == "indeterminado":
        return "Indeterminado"
    return "Sem dado"


def short_credit(value: str) -> str:
    text = normalize_text(value).lower()
    if text == "alto_potencial":
        return "Alto"
    if text == "potencial_condicionado":
        return "Condicionado"
    if text == "baixo_ou_inexistente":
        return "Baixo"
    if text == "indeterminado":
        return "Indeterminado"
    return "Sem dado"


def internal_status_label(value: str) -> str:
    text = normalize_text_upper(value)
    if text == "A":
        return "Ativa"
    if text == "I":
        return "Inativa"
    return text or "Sem dado"


def official_status_label(status: str, status_api: str) -> str:
    status_text = normalize_text(status).lower()
    api_text = normalize_text(status_api).lower()
    if status_text == "ativa":
        return "Ativa"
    if status_text and status_text not in {"nao_consultado", "não consultado"}:
        return "Nao ativa"
    if api_text and api_text not in {"api", "cache"}:
        return "Nao consultado"
    return "Sem dado"


def yes_no_label(value: str, yes: str, no: str) -> str:
    text = normalize_text(value).upper()
    if text == "S":
        return yes
    if text == "N":
        return no
    return "Sem analise"


def split_items(value: str) -> list[str]:
    return [item for item in normalize_text(value).split(";") if item]


def build_action(row: dict[str, str], has_new_data: bool, any_divergence: bool) -> str:
    if any_divergence:
        return "Revisar antes"
    if row["necessita_validacao_fiscal"] == "S":
        return "Validar fiscal"
    if row["necessita_saneamento_cadastral"] == "S":
        return "Sanear cadastro"
    if has_new_data:
        return "Atualizar BD"
    return "Sem alteracao"


def build_change_summary(row: dict[str, str]) -> tuple[list[str], bool, bool]:
    added: list[str] = []
    changed: list[str] = []

    if row["razao_social_oficial"]:
        added.append("razao oficial")
    if row["nome_fantasia_oficial"]:
        added.append("fantasia oficial")
    if row["email_oficial"]:
        added.append("email oficial")
    if row["situacao_cadastral_oficial"]:
        added.append("situacao oficial")
    if row["regime_indiciado"]:
        added.append("regime")
    if row["potencial_credito_2027"]:
        added.append("credito")
    if row["matriz_filial"]:
        added.append("matriz/filial")
    if row["porte_empresa"]:
        added.append("porte")
    if row["natureza_juridica"]:
        added.append("natureza")

    if row["divergencia_razao_social"] == "S":
        changed.append("razao")
    if row["divergencia_nome_fantasia"] == "S":
        changed.append("fantasia")
    if row["divergencia_email"] == "S":
        changed.append("email")

    summary: list[str] = []
    if added:
        summary.append(f"{len(added)} incluidos")
    if changed:
        summary.append(f"{len(changed)} divergencias")
    if not summary:
        summary.append("sem dado novo")
    return summary, bool(added), bool(changed)


def consolidate_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    company_order = {"IDEAL": 1, "MELHOR": 2, "POMME": 3}
    groups: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        key = row.get("chave") or row.get("documento") or f"{row['empresa']}::{row['codigo']}"
        groups.setdefault(key, []).append(row)

    consolidated: list[dict[str, Any]] = []
    for group_rows in groups.values():
        base = max(
            group_rows,
            key=lambda item: (
                item["sort_curva_valor"],
                item["sort_nfe_valor"],
                item["dado_novo_label"] == "Com dado novo",
            ),
        ).copy()
        empresas = sorted({row["empresa"] for row in group_rows}, key=lambda item: company_order.get(item, 99))
        codigos = sorted({row["codigo"] for row in group_rows if row["codigo"]})
        ocorrencias = [
            {
                "empresa": row["empresa"],
                "codigo": row["codigo"],
                "linha_original": row["linha_original"],
                "razao_original": row["razao_original"],
                "fantasia_original": row["fantasia_original"],
                "status_interno": row["status_interno"],
            }
            for row in sorted(group_rows, key=lambda item: (company_order.get(item["empresa"], 99), int(item["linha_original"])))
        ]

        base["empresas_lista"] = empresas
        base["empresas_label"] = ";".join(empresas)
        base["codigos_lista"] = codigos
        base["codigos_label"] = ";".join(codigos)
        base["ocorrencias"] = ocorrencias
        base["documento_formatado"] = format_document(base["documento"])
        base["row_key"] = base.get("chave") or base.get("documento") or f"{base['empresa']}-{base['linha_original']}"
        consolidated.append(base)

    return consolidated


def build_rows() -> list[dict[str, Any]]:
    normalized_map = load_csv_index(
        NORMALIZED_FILE,
        ["empresa_origem", "codigo_fornecedor", "cnpj", "razao_social"],
    )
    reconciled_by_key = load_csv_by_field(RECONCILED_FILE, "chave_unificacao")
    classified_by_key = load_csv_by_field(CLASSIFIED_FILE, "chave_unificacao")
    enriched_by_key = load_csv_by_field(ENRICHED_FILE, "chave_unificacao")

    workbook = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        iterator = ws.iter_rows(values_only=True)
        header = next(iterator)
        idx = {name: pos for pos, name in enumerate(header)}

        for excel_row_idx, row in enumerate(iterator, start=2):
            codigo = normalize_text_upper(row[idx["CDFORNECED"]])
            cnpj = normalize_doc(row[idx["NRINSJURFORN"]], 14)
            razao = normalize_text_upper(row[idx["NMRAZSOCFORN"]])
            key = (sheet_name, codigo, cnpj, razao)

            normalized = normalized_map.get(key, {})
            chave = normalized.get("chave_unificacao", "")
            reconciled = reconciled_by_key.get(chave, {})
            classified = classified_by_key.get(chave, {})
            enriched = enriched_by_key.get(chave, {})

            base = {
                "linha_original": str(excel_row_idx),
                "empresa": sheet_name,
                "codigo": codigo,
                "tipo": classify_code_type(codigo),
                "documento": cnpj or normalize_doc(codigo),
                "chave": chave,
                "razao_original": razao,
                "fantasia_original": normalize_text_upper(row[idx["NMFANTFORN"]]),
                "status_interno": normalize_text_upper(row[idx["IDSITUCADA"]]),
                "emails_originais": normalized.get("emails", ""),
                "ie_original": normalized.get("inscricao_estadual", ""),
                "im_original": normalized.get("inscricao_municipal", ""),
                "razao_social_oficial": classified.get("razao_social_oficial", ""),
                "nome_fantasia_oficial": enriched.get("nome_fantasia_oficial", ""),
                "situacao_cadastral_oficial": classified.get("situacao_cadastral_oficial", ""),
                "email_oficial": enriched.get("email_oficial", ""),
                "matriz_filial": classified.get("matriz_filial", ""),
                "porte_empresa": classified.get("porte_empresa", ""),
                "natureza_juridica": classified.get("natureza_juridica", ""),
                "opcao_simples": classified.get("opcao_simples", ""),
                "opcao_mei": classified.get("opcao_mei", ""),
                "status_api": classified.get("status_api", enriched.get("status_api", "")),
                "regime_indiciado": classified.get("regime_indiciado", ""),
                "potencial_credito_2027": classified.get("potencial_credito_2027_inicial", ""),
                "status_cadastro_oficial": classified.get("status_cadastro_oficial", ""),
                "status_regime_fiscal": classified.get("status_regime_fiscal", ""),
                "necessita_saneamento_cadastral": classified.get("necessita_saneamento_cadastral", ""),
                "motivos_saneamento_cadastral": classified.get("motivos_saneamento_cadastral", ""),
                "necessita_validacao_fiscal": classified.get("necessita_validacao_fiscal", ""),
                "motivos_validacao_fiscal": classified.get("motivos_validacao_fiscal", ""),
                "divergencia_razao_social": classified.get("divergencia_razao_social_oficial", reconciled.get("divergencia_razao_social", "")),
                "divergencia_nome_fantasia": classified.get("divergencia_nome_fantasia_oficial", reconciled.get("divergencia_nome_fantasia", "")),
                "divergencia_email": classified.get("divergencia_email_oficial", ""),
                "movimento_nfe": reconciled.get("tem_movimento_nfe", ""),
                "nfe_linhas": reconciled.get("nfe_linhas", ""),
                "nfe_valor_total": reconciled.get("nfe_valor_total", ""),
                "nfe_produtos_unicos": reconciled.get("nfe_produtos_unicos", ""),
                "nfe_ufs": reconciled.get("nfe_ufs", ""),
                "nfe_anos": reconciled.get("nfe_anos", ""),
                "curva": reconciled.get("curva_classificacao", ""),
                "curva_posicao": reconciled.get("curva_posicao", ""),
                "curva_total_fornecedor": reconciled.get("curva_total_fornecedor", ""),
                "curva_percentual": reconciled.get("curva_percentual", ""),
                "score_cadastral": reconciled.get("score_completude", normalized.get("score_completude", "")),
                "faixa_completude": reconciled.get("faixa_completude", normalized.get("faixa_completude", "")),
                "pendencias_essenciais": normalized.get("pendencias_essenciais", ""),
                "pendencias_complementares": normalized.get("pendencias_complementares", ""),
                "prioridade_saneamento_score": reconciled.get("prioridade_saneamento_score", ""),
                "prioridade_saneamento_faixa": reconciled.get("prioridade_saneamento_faixa", ""),
            }

            summary, has_new_data, any_divergence = build_change_summary(base)
            status_label = official_status_label(base["situacao_cadastral_oficial"], base["status_api"])
            action = build_action(base, has_new_data, any_divergence)
            curve_value = as_float(base["curva_total_fornecedor"])

            base.update(
                {
                    "nome_exibicao": base["razao_original"] or base["razao_social_oficial"] or base["codigo"],
                    "status_interno_label": internal_status_label(base["status_interno"]),
                    "situacao_label": status_label,
                    "regime_label": short_regime(base["regime_indiciado"]),
                    "credito_label": short_credit(base["potencial_credito_2027"]),
                    "movimento_label": "Com movimento" if base["movimento_nfe"] == "S" else "Sem movimento",
                    "curva_label": base["curva"] or "Sem curva",
                    "curva_valor_fmt": money(base["curva_total_fornecedor"]),
                    "curva_pct_fmt": pct(base["curva_percentual"]),
                    "nfe_valor_fmt": money(base["nfe_valor_total"]),
                    "dado_novo_label": "Com dado novo" if has_new_data else "Sem dado novo",
                    "divergencia_label": "Qualquer divergencia" if any_divergence else "Sem divergencia",
                    "saneamento_label": yes_no_label(base["necessita_saneamento_cadastral"], "Precisa sanear", "Nao precisa"),
                    "validacao_label": yes_no_label(base["necessita_validacao_fiscal"], "Precisa validar", "Nao precisa"),
                    "acao_label": action,
                    "mudanca_resumo": " / ".join(summary),
                    "sort_curva_rank": CURVE_ORDER.get(base["curva"], 99),
                    "sort_curva_valor": curve_value,
                    "sort_nfe_valor": as_float(base["nfe_valor_total"]),
                    "row_tone": action_to_tone(action, has_new_data),
                }
            )
            rows.append(base)

    rows = consolidate_rows(rows)
    rows.sort(
        key=lambda item: (
            item["sort_curva_rank"],
            -item["sort_curva_valor"],
            -item["sort_nfe_valor"],
            item["nome_exibicao"],
        )
    )
    return rows


def action_to_tone(action: str, has_new_data: bool) -> str:
    if action == "Revisar antes":
        return "review"
    if action == "Validar fiscal":
        return "fiscal"
    if action == "Sanear cadastro":
        return "sanitize"
    if action == "Atualizar BD":
        return "update"
    if has_new_data:
        return "update"
    return "none"


def build_html(rows: list[dict[str, Any]]) -> None:
    payload = json.dumps(rows, ensure_ascii=False)
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<title>Painel de Auditoria de Fornecedores</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
*{{box-sizing:border-box}}
body{{margin:0;background:#f3f5f8;color:#172033;font-family:Segoe UI,Arial,sans-serif;font-size:13px}}
button,select,input{{font:inherit}}
.page{{min-height:100vh;padding:14px}}
.sticky-panel{{position:sticky;top:0;z-index:12;background:#f3f5f8;padding-top:1px;padding-bottom:4px}}
.topbar{{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:10px}}
h1{{margin:0;font-size:24px;line-height:1.1;color:#111827}}
.top-actions{{display:flex;gap:8px;align-items:center}}
.icon-btn{{width:34px;height:34px;border:1px solid #cbd5e1;border-radius:8px;background:#fff;color:#111827;font-weight:800;cursor:pointer}}
.text-btn{{height:34px;border:1px solid #cbd5e1;border-radius:8px;background:#fff;color:#111827;font-weight:700;padding:0 10px;cursor:pointer}}
.filters{{display:grid;grid-template-columns:repeat(8,minmax(0,1fr));gap:7px;margin-bottom:10px}}
.filter{{min-width:0}}
.filter label{{display:flex;align-items:center;gap:4px;margin:0 0 2px;color:#64748b;font-size:10px;font-weight:700;text-transform:uppercase;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.filter.active label{{color:#1d4ed8}}
.filter input,.filter select{{width:100%;height:31px;border:1px solid #cbd5e1;border-radius:7px;background:#fff;padding:4px 7px;font-size:12px;min-width:0}}
.multi{{position:relative}}
.multi-button{{width:100%;height:31px;border:1px solid #cbd5e1;border-radius:7px;background:#fff;padding:4px 7px;font-size:12px;text-align:left;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;cursor:pointer}}
.filter.active .multi-button,.filter.active input{{border-color:#2563eb;background:#eff6ff}}
.multi-panel{{display:none;position:absolute;left:0;right:0;top:35px;z-index:15;background:#fff;border:1px solid #cbd5e1;border-radius:8px;box-shadow:0 12px 28px rgba(15,23,42,.18);max-height:240px;overflow:auto;padding:6px}}
.multi.open .multi-panel{{display:block}}
.multi-option{{display:flex;align-items:center;gap:6px;padding:5px 4px;font-size:12px;cursor:pointer}}
.multi-option input{{width:auto;height:auto;margin:0}}
.help{{display:inline-flex;align-items:center;justify-content:center;width:14px;height:14px;border:1px solid #cbd5e1;border-radius:999px;background:#fff;color:#475569;font-size:10px;font-weight:800;text-transform:none;position:relative;flex:0 0 auto}}
.help::after{{content:attr(data-tip);display:none;position:absolute;left:0;top:18px;z-index:30;width:380px;padding:10px 11px;border:1px solid #cbd5e1;border-radius:8px;background:#111827;color:#fff;font-size:11px;font-weight:500;line-height:1.45;text-transform:none;white-space:normal;box-shadow:0 12px 28px rgba(15,23,42,.25)}}
.help:hover::after{{display:block}}
.stats{{display:grid;grid-template-columns:repeat(6,minmax(0,1fr));gap:7px;margin-bottom:10px}}
.stat{{background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:8px 10px;min-width:0}}
.stat .label{{font-size:10px;color:#64748b;font-weight:700;text-transform:uppercase;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.stat .value{{font-size:20px;font-weight:800;margin-top:3px;color:#111827}}
.pager{{display:grid;grid-template-columns:26px 26px minmax(42px,1fr) 26px 26px 56px;gap:4px;align-items:center;margin-top:4px}}
.pager button{{height:25px;border:1px solid #cbd5e1;border-radius:6px;background:#fff;cursor:pointer;font-weight:800;line-height:1}}
.pager input,.pager select{{height:25px;border:1px solid #cbd5e1;border-radius:6px;background:#fff;text-align:center;font-size:12px;min-width:0}}
.pager input{{width:100%}}
.table{{width:100%;border-collapse:collapse;table-layout:fixed;background:#fff;border:1px solid #dbe1ea;border-radius:8px;overflow:hidden}}
th{{height:34px;text-align:left;padding:7px 8px;background:#e8edf5;color:#172033;font-size:11px;text-transform:uppercase;border-bottom:1px solid #dbe1ea;position:sticky;top:var(--sticky-offset,0px);z-index:11}}
td{{padding:8px;border-bottom:1px solid #e5e7eb;vertical-align:middle;overflow:hidden}}
tr.summary{{cursor:pointer}}
tr.summary:hover{{background:#f8fafc}}
tr.tone-review td:first-child{{border-left:5px solid #f59e0b}}
tr.tone-fiscal td:first-child{{border-left:5px solid #ef4444}}
tr.tone-sanitize td:first-child{{border-left:5px solid #f97316}}
tr.tone-update td:first-child{{border-left:5px solid #2563eb}}
tr.tone-none td:first-child{{border-left:5px solid #cbd5e1}}
.col-fornecedor{{width:26%}}
.col-mudanca{{width:16%}}
.col-fiscal{{width:13%}}
.col-credito{{width:8%}}
.col-movimento{{width:14%}}
.col-acao{{width:12%}}
.col-prioridade{{width:11%}}
.name{{font-weight:800;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.sub{{color:#64748b;font-size:11px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-top:3px}}
.supplier-meta{{display:flex;align-items:center;gap:5px;min-width:0;margin-top:3px;white-space:nowrap;overflow:hidden}}
.company-tag{{display:inline-flex;align-items:center;height:18px;padding:0 6px;border-radius:5px;font-size:10px;font-weight:800;color:#fff;flex:0 0 auto}}
.company-tag.ideal{{background:#2563eb}}
.company-tag.melhor{{background:#16a34a}}
.company-tag.pomme{{background:#dc2626}}
.doc{{color:#475569;font-size:11px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.pill{{display:inline-flex;align-items:center;max-width:100%;height:22px;padding:0 8px;border-radius:999px;font-size:11px;font-weight:800;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.pill.high{{background:#dcfce7;color:#166534}}
.pill.mid{{background:#fef3c7;color:#92400e}}
.pill.risk{{background:#fee2e2;color:#991b1b}}
.pill.info{{background:#dbeafe;color:#1d4ed8}}
.pill.gray{{background:#e5e7eb;color:#334155}}
.pill.review{{background:#ffedd5;color:#9a3412}}
.line-tight{{white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.detail-row{{display:none;background:#fbfcfe}}
.detail-row.open{{display:table-row}}
.detail-cell{{padding:0 8px 10px 8px}}
.detail-grid{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:8px;padding:9px;border-top:1px solid #e5e7eb}}
.panel{{background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:9px;min-width:0}}
.panel h3{{margin:0 0 7px;font-size:12px;color:#111827}}
.kv{{display:grid;grid-template-columns:96px minmax(0,1fr);gap:5px;margin:3px 0}}
.kv b{{font-size:11px;color:#475569}}
.kv span{{min-width:0;overflow-wrap:anywhere}}
.tags{{display:flex;flex-wrap:wrap;gap:5px}}
.tag{{display:inline-flex;max-width:100%;padding:4px 7px;border-radius:7px;font-size:11px;font-weight:700;background:#e5e7eb;color:#334155}}
.tag.blue{{background:#dbeafe;color:#1d4ed8}}
.tag.yellow{{background:#fef3c7;color:#92400e}}
.tag.red{{background:#fee2e2;color:#991b1b}}
.tag.green{{background:#dcfce7;color:#166534}}
.modal-backdrop{{position:fixed;inset:0;background:rgba(15,23,42,.36);display:none;align-items:center;justify-content:center;padding:20px;z-index:20}}
.modal-backdrop.open{{display:flex}}
.modal{{width:min(880px,100%);max-height:88vh;overflow:auto;background:#fff;border-radius:10px;box-shadow:0 24px 70px rgba(15,23,42,.35);padding:18px}}
.modal-header{{display:flex;justify-content:space-between;align-items:center;gap:12px;margin-bottom:10px}}
.modal h2{{margin:0;font-size:20px}}
.modal h3{{margin:14px 0 6px;font-size:14px}}
.modal p,.modal li{{line-height:1.45}}
.field-grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:8px}}
.field-grid label{{display:flex;gap:7px;align-items:center}}
.empty{{padding:24px;text-align:center;color:#64748b;background:#fff;border:1px dashed #cbd5e1;border-radius:8px}}
@media (max-width: 1280px){{
  .filters{{grid-template-columns:repeat(4,minmax(0,1fr))}}
  .stats{{grid-template-columns:repeat(3,minmax(0,1fr))}}
  .detail-grid{{grid-template-columns:repeat(2,minmax(0,1fr))}}
}}
@media (max-width: 760px){{
  .page{{padding:10px}}
  .filters,.stats,.detail-grid{{grid-template-columns:1fr}}
  .table,.table thead,.table tbody,.table tr,.table th,.table td{{display:block;width:100%}}
  .table thead{{display:none}}
  tr.summary{{border-bottom:1px solid #dbe1ea}}
  td[data-label]::before{{content:attr(data-label);display:block;font-size:10px;color:#64748b;font-weight:800;text-transform:uppercase;margin-bottom:3px}}
  .detail-row.open{{display:block}}
}}
</style>
</head>
<body>
<div class="page">
  <div class="sticky-panel" id="stickyPanel">
  <div class="topbar">
    <h1>Painel de Fornecedores</h1>
    <div class="top-actions">
      <button class="icon-btn" id="methodBtn" title="Metodologia">?</button>
    </div>
  </div>

  <div class="filters" id="filters">
    <div class="filter"><label>Busca <span class="help" data-tip="Procura por codigo interno, CNPJ/CPF, razao social, nome fantasia, email e acao sugerida.">?</span></label><input id="q" placeholder="codigo, CNPJ, nome, email"></div>
    <div class="filter"><label>Empresa <span class="help" data-tip="Filtra fornecedores presentes em IDEAL, MELHOR e/ou POMME. As cores das tags seguem esta mesma regra.">?</span></label><select id="empresa"></select></div>
    <div class="filter"><label>Tipo <span class="help" data-tip="PJ/CNPJ vem de codigos J, PF/CPF vem de codigos F, e Outro cobre excecoes operacionais.">?</span></label><select id="tipo"></select></div>
    <div class="filter"><label>Movimento NFe <span class="help" data-tip="Indica se o fornecedor tem notas fiscais de entrada no historico analisado.">?</span></label><select id="movimento"></select></div>
    <div class="filter"><label>Curva <span class="help" data-tip="Classificacao da curva de fornecedor por valor comprado: AAA maior relevancia, CCC menor relevancia, Sem curva quando nao aparece na curva.">?</span></label><select id="curva"></select></div>
    <div class="filter"><label>Situação oficial <span class="help" data-tip="Situacao cadastral retornada pela OpenCNPJ ou ausencia de consulta/dado oficial.">?</span></label><select id="situacao"></select></div>
    <div class="filter"><label>Regime <span class="help" data-tip="Leitura inicial do regime: Simples, Nao Simples, MEI, Indeterminado ou Sem dado.">?</span></label><select id="regime"></select></div>
    <div class="filter"><label>Crédito <span class="help" data-tip="Potencial inicial de credito em 2027: Alto, Condicionado, Baixo, Indeterminado ou Sem dado.">?</span></label><select id="credito"></select></div>
    <div class="filter"><label>Dado novo <span class="help" data-tip="Mostra se a etapa de enriquecimento acrescentou informacoes oficiais ou fiscais ao fornecedor.">?</span></label><select id="dadoNovo"></select></div>
    <div class="filter"><label>Divergência <span class="help" data-tip="Compara cadastro interno com dados oficiais em razao social, fantasia e email.">?</span></label><select id="divergencia"></select></div>
    <div class="filter"><label>Saneamento <span class="help" data-tip="Indica se faltam dados cadastrais essenciais para o fornecedor.">?</span></label><select id="saneamento"></select></div>
    <div class="filter"><label>Validação fiscal <span class="help" data-tip="Indica se a leitura fiscal precisa de revisao por regime indefinido, divergencia ou outro alerta.">?</span></label><select id="validacao"></select></div>
    <div class="filter"><label>Ação <span class="help" data-tip="Acao sugerida: atualizar BD, revisar antes, validar fiscal, sanear cadastro ou sem alteracao.">?</span></label><select id="acao"></select></div>
    <div class="filter"><label>Completude <span class="help" data-tip="Faixa do score cadastral calculado na fase 1: alta, media ou baixa.">?</span></label><select id="completude"></select></div>
    <div class="filter"><label>Prioridade <span class="help" data-tip="Prioridade de saneamento considerando movimento, curva, pendencias e divergencias.">?</span></label><select id="prioridade"></select></div>
    <div class="filter"><label>Campos <span class="help" data-tip="Escolhe quais colunas resumidas aparecem na tabela principal.">?</span></label><select id="campos"></select></div>
  </div>

  <div class="stats">
    <div class="stat"><div class="label">Registros</div><div class="value" id="stRows">0</div></div>
    <div class="stat"><div class="label">Com dado novo</div><div class="value" id="stNew">0</div></div>
    <div class="stat"><div class="label">Com movimento</div><div class="value" id="stMov">0</div></div>
    <div class="stat"><div class="label">Revisar antes</div><div class="value" id="stReview">0</div></div>
    <div class="stat"><div class="label">Validar fiscal</div><div class="value" id="stFiscal">0</div></div>
    <div class="stat"><div class="label">Paginação</div>
      <div class="pager">
        <button id="firstPage" title="Primeira página">⏮</button>
        <button id="prevPage" title="Página anterior">◀</button>
        <input id="pageInput" value="1" aria-label="Número da página">
        <button id="nextPage" title="Próxima página">▶</button>
        <button id="lastPage" title="Última página">⏭</button>
        <select id="pageSize" aria-label="Itens por página">
          <option>10</option><option selected>20</option><option>30</option><option>40</option><option>50</option><option>100</option>
        </select>
      </div>
    </div>
  </div>
  </div>

  <table class="table">
    <thead>
      <tr>
        <th class="col-fornecedor" data-col="fornecedor">Fornecedor <span class="help" data-tip="Nome do fornecedor consolidado, tags das empresas onde aparece e CNPJ/CPF formatado.">?</span></th>
        <th class="col-mudanca" data-col="mudanca">Mudou / Incluiu <span class="help" data-tip="Resumo do que foi acrescentado pela consulta externa e das divergencias encontradas.">?</span></th>
        <th class="col-fiscal" data-col="fiscal">Fiscal <span class="help" data-tip="Situacao oficial e regime fiscal indicado.">?</span></th>
        <th class="col-credito" data-col="credito">Crédito <span class="help" data-tip="Potencial inicial de credito fiscal para 2027: Alto, Condicionado, Baixo, Indeterminado ou Sem dado.">?</span></th>
        <th class="col-movimento" data-col="movimento">Movimento <span class="help" data-tip="Curva do fornecedor, valor da curva e valor total de NFe identificado.">?</span></th>
        <th class="col-acao" data-col="acao">Ação <span class="help" data-tip="Proxima acao sugerida para uso operacional do cadastro.">?</span></th>
        <th class="col-prioridade" data-col="prioridade">Score <span class="help" data-tip="Score e faixa de completude cadastral, mais prioridade de saneamento.">?</span></th>
      </tr>
    </thead>
    <tbody id="tbody"></tbody>
  </table>
  <div class="empty" id="empty" style="display:none">Nenhum fornecedor encontrado com os filtros atuais.</div>
</div>

<div class="modal-backdrop" id="methodModal">
  <div class="modal">
    <div class="modal-header">
      <h2>Metodologia</h2>
      <button class="icon-btn" data-close="methodModal">x</button>
    </div>
    <h3>Objetivo da tela</h3>
    <p>Este painel foi pensado como uma tela de auditoria cadastral e fiscal. A linha principal mostra uma versao consolidada do fornecedor para leitura rapida. O detalhe expandido preserva as ocorrencias originais por empresa, os dados internos, o retorno oficial e os motivos de revisao.</p>
    <h3>Como a base foi montada</h3>
    <p>A planilha original continua sendo a fonte de verdade do cadastro. Cada linha original foi lida com empresa de origem, codigo interno, documento, razao social, fantasia, emails, inscricoes e situacao interna. Em seguida, os registros foram agrupados por chave de fornecedor, principalmente CNPJ quando disponivel.</p>
    <p>Quando o mesmo fornecedor aparece em mais de uma empresa do grupo, ele aparece uma unica vez na lista principal. As tags coloridas mostram onde ele existe: IDEAL em azul, MELHOR em verde e POMME em vermelho. As linhas originais e codigos internos continuam guardados no detalhe, para rastreabilidade e eventual devolucao ao BD.</p>
    <h3>Movimento e curva</h3>
    <p>O painel cruza o cadastro com NFe e curva de fornecedor. A ordenacao inicial prioriza a curva por valor: AAA primeiro, depois AA, A, B, BB, C, CC e CCC. Dentro de cada faixa, vem primeiro quem tem maior valor de compra na curva.</p>
    <p>A coluna Movimento mostra a curva, o valor da curva e o valor de NFe localizado. Isso ajuda a separar fornecedor importante de fornecedor apenas cadastrado.</p>
    <h3>Enriquecimento OpenCNPJ</h3>
    <p>Para fornecedores PJ/CNPJ priorizados, foi consultada a OpenCNPJ. O painel pode acrescentar razao social oficial, nome fantasia oficial, situacao cadastral, matriz/filial, porte, natureza juridica, opcao pelo Simples, opcao pelo MEI e email oficial.</p>
    <p>Quando o cadastro interno difere do retorno oficial, a divergencia aparece como alerta. Divergencia nao significa que o cadastro interno esta necessariamente errado; significa que a diferenca precisa ser revisada antes de atualizar o banco.</p>
    <h3>Leitura fiscal 2026 e 2027</h3>
    <p>Em 2026, a Reforma Tributaria do Consumo esta em fase de teste operacional. Os documentos passam a conviver com informacoes de CBS e IBS, mas a leitura de credito ainda deve ser tratada com cuidado.</p>
    <p>A partir de 2027, a CBS passa a ter efeito material e o regime do fornecedor passa a influenciar a expectativa de credito para a empresa compradora. Por isso, este painel nao olha apenas preco ou volume: ele tambem marca regime, Simples, MEI, indeterminacao fiscal e necessidade de validacao.</p>
    <h3>Credito</h3>
    <p><b>Alto</b>: fornecedor com indicio de maior potencial de credito. <b>Condicionado</b>: potencial depende de condicoes fiscais, especialmente em fornecedores do Simples. <b>Baixo</b>: potencial baixo ou inexistente, como em casos de MEI. <b>Indeterminado</b>: faltam dados suficientes para classificar. <b>Sem dado</b>: fornecedor nao recebeu analise fiscal nesta fase.</p>
    <h3>Acoes sugeridas</h3>
    <p><b>Atualizar BD</b>: ha dado novo sem alerta relevante. <b>Revisar antes</b>: existe divergencia entre dado interno e dado oficial. <b>Validar fiscal</b>: o ponto fiscal ainda precisa de revisao. <b>Sanear cadastro</b>: faltam campos cadastrais essenciais. <b>Sem alteracao</b>: nao houve enriquecimento util nesta fase.</p>
    <h3>Limites da analise</h3>
    <p>Esta etapa e uma triagem operacional. Ela organiza evidencias, aponta diferencas e cria prioridade. A classificacao fiscal ainda deve ser validada pela area fiscal antes de virar regra definitiva de credito ou de atualizacao automatica do cadastro.</p>
    <h3>Fontes usadas e papel de cada uma</h3>
    <p>A planilha de fornecedores e a base de origem do cadastro. Ela representa aquilo que hoje existe no ambiente operacional das empresas do grupo e, por isso, e tratada como ponto de partida e nao como dado automaticamente correto. A base NFE entra para dizer se o fornecedor teve movimento real de compras, em quais anos, em quais UFs e com qual valor acumulado. A curva de fornecedores entra para dar prioridade economica, porque um cadastro incompleto em fornecedor AAA tem risco operacional muito diferente de um cadastro incompleto sem movimento.</p>
    <p>A consulta OpenCNPJ entra como fonte externa de enriquecimento e validacao inicial para pessoas juridicas. Ela nao substitui validacao fiscal formal, mas ajuda a preencher campos oficiais, identificar situacao cadastral, natureza juridica, porte, opcao pelo Simples Nacional, opcao pelo MEI e dados de contato quando disponiveis.</p>
    <h3>Consolidacao por fornecedor</h3>
    <p>O cadastro original possui linhas por empresa e por codigo interno. Como um mesmo CNPJ pode aparecer em IDEAL, MELHOR e POMME, o painel consolida fornecedores por documento/chave para evitar leitura duplicada. A lista principal mostra um unico fornecedor, enquanto as tags de empresa indicam onde ele aparece. Isso reduz ruido visual e permite comparar risco, movimento e credito por entidade economica, nao por repeticao administrativa.</p>
    <p>A consolidacao nao apaga a rastreabilidade. Os codigos internos, linhas de origem e ocorrencias por empresa continuam carregados nos dados do painel, mesmo quando nao aparecem como tags visuais no detalhe. Essa decisao separa auditoria visual de trilha tecnica: a tela fica legivel, mas o dado de origem continua disponivel para retorno ao banco ou conciliacao posterior.</p>
    <h3>Qualidade cadastral</h3>
    <p>A completude cadastral foi calculada como indicador operacional. Ela considera presenca de documento, razao social, fantasia, email, inscricoes, situacao interna e outros campos relevantes da estrutura cadastral. O objetivo do score nao e dizer que o fornecedor esta fiscalmente regular; o objetivo e mostrar se o cadastro tem informacao suficiente para operar, auditar e atualizar com seguranca.</p>
    <p>As pendencias essenciais aparecem no box Cadastro original quando correspondem a campos exibidos ali, como email e inscricoes. Pendencias complementares como telefone de suporte, site e grupo fiscal nao sao mais mostradas no box Operacao e saneamento, porque nao pertencem a curva, NFe, UF ou anos. Elas continuam existindo na base de dados, mas nao entram como tag em um box onde poderiam induzir leitura errada.</p>
    <h3>Divergencias cadastrais</h3>
    <p>Divergencia e diferenca entre o que esta no cadastro interno e o que veio da fonte externa. Divergencia de razao social, fantasia ou email nao significa automaticamente erro do cadastro interno nem erro da fonte externa. Significa que atualizar automaticamente seria arriscado. Por isso, fornecedores com divergencia tendem a receber acao Revisar antes, especialmente quando tambem possuem movimento ou relevancia na curva.</p>
    <h3>Movimento, curva e prioridade</h3>
    <p>Movimento NFe indica fornecedor que teve nota fiscal de entrada localizada. Curva indica peso economico por valor comprado. A prioridade combina esses sinais com qualidade cadastral e risco fiscal. Um fornecedor AAA com divergencia ou regime indeterminado deve aparecer antes de um fornecedor sem curva, porque o impacto de uma decisao errada e maior.</p>
    <p>O painel foi ordenado inicialmente por curva de fornecedor por valor: AAA, AA, A, B, BB, C, CC, CCC e, depois, fornecedores sem curva. Dentro de cada faixa, fornecedores com maior valor aparecem primeiro. Essa ordenacao e intencional para transformar o painel em fila de trabalho, nao apenas em relatorio.</p>
    <h3>Leitura fiscal e Reforma Tributaria</h3>
    <p>A leitura fiscal foi desenhada considerando a transicao 2026/2027. Em 2026, a Reforma Tributaria do Consumo entra em ambiente de teste e convivencia informacional de CBS e IBS. Em 2027, a CBS passa a ter impacto material e substitui PIS/Cofins, tornando a qualidade fiscal do fornecedor mais importante para a tomada de decisao de compras e para a expectativa de credito.</p>
    <p>O painel nao calcula credito definitivo por produto. Ele classifica o fornecedor como ponto de partida. A geracao efetiva de credito dependera tambem da natureza da operacao, do item comprado, da classificacao tributaria, do documento fiscal, do regime do fornecedor e de regras especificas que ainda precisam ser confirmadas pela area fiscal. Portanto, Credito no painel e uma tag de triagem, nao uma apuracao tributaria final.</p>
    <h3>Regime e credito</h3>
    <p>Fornecedor Nao Simples tende a receber leitura de Alto potencial quando ha dado suficiente e situacao cadastral favoravel, porque em tese tende a operar fora do regime simplificado. Fornecedor do Simples recebe Condicionado porque o impacto de credito para o adquirente pode depender da forma de apuracao e de regras especificas aplicaveis. MEI tende a Baixo por restricoes e baixa expectativa de credito. Indeterminado exige validacao fiscal antes de qualquer decisao operacional. Sem dado significa que a analise externa ainda nao trouxe informacao aproveitavel.</p>
    <h3>Acoes operacionais</h3>
    <p>Atualizar BD significa que ha dado novo util e sem alerta forte de divergencia. Revisar antes significa que existe conflito entre cadastro interno e fonte externa, entao a atualizacao automatica seria perigosa. Validar fiscal significa que a situacao tributaria ou o regime nao estao suficientemente claros. Sanear cadastro significa que faltam campos internos essenciais. Sem alteracao significa que a fase atual nao encontrou ganho pratico para aquele fornecedor.</p>
    <h3>Como ler as cores e tags</h3>
    <p>As cores nao sao decorativas; elas indicam tipo de atencao. Verde tende a indicar dado favoravel ou ausencia de alerta principal. Azul indica informacao relevante ou impacto economico. Amarelo indica pendencia que merece saneamento. Vermelho indica risco, divergencia ou necessidade de validacao. Cinza indica falta de dado ou situacao neutra.</p>
    <h3>Como usar o filtro Todos</h3>
    <p>Nos filtros de multipla escolha, Todos e o estado neutro. Quando Todos esta marcado, todas as opcoes internas ficam marcadas e o filtro nao restringe a base. Se uma opcao especifica for desmarcada, Todos deixa de ser o estado ativo e o filtro passa a restringir. Se nenhuma opcao ficar marcada, o filtro retorna zero resultados, o que representa uma exclusao total intencional.</p>
  </div>
</div>

<script>
const DATA = {payload};
const state = {{ expandedKey: '', page: 1, pageSize: 20, multi: {{}} }};
const els = {{
  q: document.getElementById('q'),
  empresa: document.getElementById('empresa'),
  tipo: document.getElementById('tipo'),
  movimento: document.getElementById('movimento'),
  curva: document.getElementById('curva'),
  situacao: document.getElementById('situacao'),
  regime: document.getElementById('regime'),
  credito: document.getElementById('credito'),
  dadoNovo: document.getElementById('dadoNovo'),
  divergencia: document.getElementById('divergencia'),
  saneamento: document.getElementById('saneamento'),
  validacao: document.getElementById('validacao'),
  acao: document.getElementById('acao'),
  completude: document.getElementById('completude'),
  prioridade: document.getElementById('prioridade'),
  campos: document.getElementById('campos'),
  tbody: document.getElementById('tbody'),
  empty: document.getElementById('empty'),
  stRows: document.getElementById('stRows'),
  stNew: document.getElementById('stNew'),
  stMov: document.getElementById('stMov'),
  stReview: document.getElementById('stReview'),
  stFiscal: document.getElementById('stFiscal'),
  firstPage: document.getElementById('firstPage'),
  prevPage: document.getElementById('prevPage'),
  pageInput: document.getElementById('pageInput'),
  nextPage: document.getElementById('nextPage'),
  lastPage: document.getElementById('lastPage'),
  pageSize: document.getElementById('pageSize'),
}};

const columns = [
  ['fornecedor', 'Fornecedor'],
  ['mudanca', 'Mudou / Incluiu'],
  ['fiscal', 'Fiscal'],
  ['credito', 'Crédito'],
  ['movimento', 'Movimento'],
  ['acao', 'Ação'],
  ['prioridade', 'Score'],
];

function money(value) {{
  const n = Number(value || 0);
  if (!n) return 'R$ 0';
  return 'R$ ' + Math.round(n).toLocaleString('pt-BR');
}}

function safe(value) {{
  return value || 'Sem dado';
}}

function pillClass(value) {{
  if (['Alto', 'Ativa', 'Atualizar BD', 'Nao precisa'].includes(value)) return 'high';
  if (['Condicionado', 'Revisar antes', 'Sanear cadastro'].includes(value)) return 'review';
  if (['Baixo', 'Validar fiscal', 'Precisa validar', 'Precisa sanear', 'Nao ativa'].includes(value)) return 'risk';
  if (['Indeterminado'].includes(value)) return 'mid';
  if (['Com dado novo'].includes(value)) return 'info';
  return 'gray';
}}

function option(el, label, values) {{
  el.innerHTML = `<option value="">${{label}}</option>` + values.map(v => `<option value="${{v}}">${{v}}</option>`).join('');
}}

function enhanceMultiSelect(el) {{
  const values = [...el.options].slice(1).map(option => option.value);
  el.style.display = 'none';
  const multi = document.createElement('div');
  multi.className = 'multi';
  multi.innerHTML = `
    <button type="button" class="multi-button" data-multi-button="${{el.id}}">Todos</button>
    <div class="multi-panel">
      <label class="multi-option">
        <input type="checkbox" value="__ALL__" checked>
        <span>Todos</span>
      </label>
      ${{values.map(value => `
        <label class="multi-option">
          <input type="checkbox" value="${{value}}" checked>
          <span>${{value}}</span>
        </label>
      `).join('')}}
    </div>`;
  el.after(multi);
  state.multi[el.id] = {{ element: el, root: multi, values: [] }};
  multi.addEventListener('click', event => event.stopPropagation());
  const button = multi.querySelector('.multi-button');
  button.addEventListener('click', event => {{
    event.stopPropagation();
    document.querySelectorAll('.multi.open').forEach(open => {{
      if (open !== multi) open.classList.remove('open');
    }});
    multi.classList.toggle('open');
  }});
  multi.querySelectorAll('input').forEach(input => {{
    input.addEventListener('change', () => {{
      const all = multi.querySelector('input[value="__ALL__"]');
      const options = [...multi.querySelectorAll('input:not([value="__ALL__"])')];
      if (input.value === '__ALL__') {{
        options.forEach(item => item.checked = input.checked);
      }} else {{
        all.checked = options.every(item => item.checked);
      }}
      syncMultiValues(el.id);
      updateMultiLabel(el.id);
      state.page = 1;
      state.expandedKey = '';
      render();
    }});
  }});
  syncMultiValues(el.id);
  updateMultiLabel(el.id);
}}

function syncMultiValues(id) {{
  const control = state.multi[id];
  if (!control) return;
  const options = [...control.root.querySelectorAll('input:not([value="__ALL__"])')];
  const checked = options.filter(item => item.checked).map(item => item.value);
  if (checked.length === options.length) control.values = [];
  else if (!checked.length) control.values = ['__NONE__'];
  else control.values = checked;
}}

function updateMultiLabel(id) {{
  const control = state.multi[id];
  if (!control) return;
  const button = control.root.querySelector('.multi-button');
  const selected = control.values;
  control.root.closest('.filter')?.classList.toggle('active', selected.length > 0);
  if (!selected.length) button.textContent = 'Todos';
  else if (selected.includes('__NONE__')) button.textContent = 'Nenhum selecionado';
  else if (selected.length === 1) button.textContent = selected[0];
  else button.textContent = `${{selected.length}} selecionados`;
}}

function selected(id) {{
  return state.multi[id]?.values || [];
}}

function selectedMatches(id, value) {{
  const values = selected(id);
  if (values.includes('__NONE__')) return false;
  return !values.length || values.includes(value);
}}

function initFilters() {{
  option(els.empresa, 'Todas', ['IDEAL', 'MELHOR', 'POMME']);
  option(els.tipo, 'Todos', ['PJ/CNPJ', 'PF/CPF', 'Outro']);
  option(els.movimento, 'Todos', ['Com movimento', 'Sem movimento']);
  option(els.curva, 'Todas', ['AAA', 'AA', 'A', 'B', 'BB', 'C', 'CC', 'CCC', 'Sem curva']);
  option(els.situacao, 'Todas', ['Ativa', 'Nao ativa', 'Nao consultado', 'Sem dado']);
  option(els.regime, 'Todos', ['Simples', 'Nao Simples', 'MEI', 'Indeterminado', 'Sem dado']);
  option(els.credito, 'Todos', ['Alto', 'Condicionado', 'Baixo', 'Indeterminado', 'Sem dado']);
  option(els.dadoNovo, 'Todos', ['Com dado novo', 'Sem dado novo']);
  option(els.divergencia, 'Todas', ['Razao social', 'Fantasia', 'Email', 'Qualquer divergencia', 'Sem divergencia']);
  option(els.saneamento, 'Todos', ['Precisa sanear', 'Nao precisa', 'Sem analise']);
  option(els.validacao, 'Todas', ['Precisa validar', 'Nao precisa', 'Sem analise']);
  option(els.acao, 'Todas', ['Atualizar BD', 'Revisar antes', 'Validar fiscal', 'Sanear cadastro', 'Sem alteracao']);
  option(els.completude, 'Todas', ['alta', 'media', 'baixa']);
  option(els.prioridade, 'Todas', ['critica', 'alta', 'media', 'baixa']);
  option(els.campos, 'Todos', columns.map(([, label]) => label));
  [
    els.empresa, els.tipo, els.movimento, els.curva, els.situacao, els.regime, els.credito,
    els.dadoNovo, els.divergencia, els.saneamento, els.validacao, els.acao, els.completude, els.prioridade, els.campos
  ].forEach(enhanceMultiSelect);
}}

function matchesDivergence(row, value) {{
  if (value.includes('__NONE__')) return false;
  if (!value.length) return true;
  return value.some(item => {{
    if (item === 'Razao social') return row.divergencia_razao_social === 'S';
    if (item === 'Fantasia') return row.divergencia_nome_fantasia === 'S';
    if (item === 'Email') return row.divergencia_email === 'S';
    if (item === 'Qualquer divergencia') return row.divergencia_label === 'Qualquer divergencia';
    return row.divergencia_label === 'Sem divergencia';
  }});
}}

function filteredRows() {{
  const q = els.q.value.trim().toLowerCase();
  return DATA.filter(row => {{
    const empresas = selected('empresa');
    if (empresas.length && !row.empresas_lista.some(empresa => empresas.includes(empresa))) return false;
    if (!selectedMatches('tipo', row.tipo)) return false;
    if (!selectedMatches('movimento', row.movimento_label)) return false;
    if (!selectedMatches('curva', row.curva_label)) return false;
    if (!selectedMatches('situacao', row.situacao_label)) return false;
    if (!selectedMatches('regime', row.regime_label)) return false;
    if (!selectedMatches('credito', row.credito_label)) return false;
    if (!selectedMatches('dadoNovo', row.dado_novo_label)) return false;
    if (!matchesDivergence(row, selected('divergencia'))) return false;
    if (!selectedMatches('saneamento', row.saneamento_label)) return false;
    if (!selectedMatches('validacao', row.validacao_label)) return false;
    if (!selectedMatches('acao', row.acao_label)) return false;
    if (!selectedMatches('completude', row.faixa_completude)) return false;
    if (!selectedMatches('prioridade', row.prioridade_saneamento_faixa)) return false;
    if (!q) return true;
    return [
      row.codigos_label, row.documento, row.documento_formatado, row.razao_original, row.fantasia_original, row.emails_originais,
      row.razao_social_oficial, row.nome_fantasia_oficial, row.email_oficial, row.acao_label
    ].join(' ').toLowerCase().includes(q);
  }});
}}

function tags(value, color) {{
  const parts = String(value || '').split(';').filter(Boolean);
  if (!parts.length) return '<span class="tag">Sem dado</span>';
  return parts.map(item => `<span class="tag ${{color}}">${{item}}</span>`).join('');
}}

function companyTags(row) {{
  return row.empresas_lista.map(empresa => `<span class="company-tag ${{empresa.toLowerCase()}}">${{empresa}}</span>`).join('');
}}

function occurrenceTags(row) {{
  return row.ocorrencias.map(item => `<span class="tag">${{item.empresa}} linha ${{item.linha_original}}: ${{item.codigo}}</span>`).join('');
}}

function cadastroTags(row) {{
  const out = [];
  const missing = String(row.pendencias_essenciais || '').split(';').filter(Boolean);
  if (missing.includes('email')) out.push('<span class="tag yellow">email faltante</span>');
  if (missing.includes('inscricao_estadual')) out.push('<span class="tag yellow">inscricao estadual faltante</span>');
  if (missing.includes('inscricao_municipal')) out.push('<span class="tag yellow">inscricao municipal faltante</span>');
  if (row.divergencia_razao_social === 'S') out.push('<span class="tag red">divergencia razao social</span>');
  if (row.divergencia_nome_fantasia === 'S') out.push('<span class="tag red">divergencia fantasia</span>');
  if (row.divergencia_email === 'S') out.push('<span class="tag red">divergencia email</span>');
  return out.join('') || '<span class="tag green">cadastro sem alerta principal</span>';
}}

function fiscalTags(row) {{
  const fiscal = String(row.motivos_validacao_fiscal || '')
    .split(';')
    .filter(item => item && !item.startsWith('divergencia_'));
  if (!fiscal.length) return '<span class="tag green">sem alerta fiscal principal</span>';
  return fiscal.map(item => `<span class="tag red">${{item}}</span>`).join('');
}}

function operacaoTags(row) {{
  const out = [];
  if (row.movimento_label === 'Sem movimento') out.push('<span class="tag yellow">sem movimento NFe</span>');
  if (row.curva_label === 'Sem curva') out.push('<span class="tag yellow">sem curva</span>');
  if (['AAA', 'AA', 'A'].includes(row.curva_label)) out.push('<span class="tag blue">alto impacto de compra</span>');
  return out.join('') || '<span class="tag green">operacao sem alerta principal</span>';
}}

function summaryRow(row) {{
  return `
    <tr class="summary tone-${{row.row_tone}}" data-key="${{row.row_key}}">
      <td data-col="fornecedor" data-label="Fornecedor"><div class="name">${{safe(row.nome_exibicao)}}</div><div class="supplier-meta">${{companyTags(row)}}<span class="doc">${{safe(row.documento_formatado)}}</span></div></td>
      <td data-col="mudanca" data-label="Mudou / Incluiu"><div class="line-tight">${{row.mudanca_resumo}}</div><div class="sub">${{row.dado_novo_label}} | ${{row.divergencia_label}}</div></td>
      <td data-col="fiscal" data-label="Fiscal"><span class="pill ${{pillClass(row.situacao_label)}}">${{row.situacao_label}}</span><div class="sub">${{row.regime_label}}</div></td>
      <td data-col="credito" data-label="Credito"><span class="pill ${{pillClass(row.credito_label)}}">${{row.credito_label}}</span></td>
      <td data-col="movimento" data-label="Movimento"><div class="line-tight">${{row.curva_label}} | ${{row.curva_valor_fmt}}</div><div class="sub">NFe ${{row.nfe_valor_fmt}}</div></td>
      <td data-col="acao" data-label="Acao"><span class="pill ${{pillClass(row.acao_label)}}">${{row.acao_label}}</span></td>
      <td data-col="prioridade" data-label="Score"><div class="line-tight">${{safe(row.score_cadastral)}} | ${{safe(row.faixa_completude)}}</div><div class="sub">${{safe(row.prioridade_saneamento_faixa)}}</div></td>
    </tr>
    <tr class="detail-row" data-detail="${{row.row_key}}">
      <td class="detail-cell" colspan="7">
        <div class="detail-grid">
          <div class="panel"><h3>Cadastro original</h3>
            <div class="kv"><b>Razao</b><span>${{safe(row.razao_original)}}</span></div>
            <div class="kv"><b>Fantasia</b><span>${{safe(row.fantasia_original)}}</span></div>
            <div class="kv"><b>Situacao</b><span>${{safe(row.status_interno_label)}}</span></div>
            <div class="kv"><b>CNPJ/CPF</b><span>${{safe(row.documento_formatado)}}</span></div>
            <div class="kv"><b>IE</b><span>${{safe(row.ie_original)}}</span></div>
            <div class="kv"><b>IM</b><span>${{safe(row.im_original)}}</span></div>
            <div class="kv"><b>Email</b><span>${{safe(row.emails_originais)}}</span></div>
            <div class="tags">${{cadastroTags(row)}}</div>
          </div>
          <div class="panel"><h3>Dados incluidos</h3>
            <div class="kv"><b>Razao oficial</b><span>${{safe(row.razao_social_oficial)}}</span></div>
            <div class="kv"><b>Fantasia oficial</b><span>${{safe(row.nome_fantasia_oficial)}}</span></div>
            <div class="kv"><b>Status oficial</b><span>${{safe(row.situacao_cadastral_oficial)}}</span></div>
            <div class="kv"><b>Email oficial</b><span>${{safe(row.email_oficial)}}</span></div>
            <div class="kv"><b>Porte</b><span>${{safe(row.porte_empresa)}}</span></div>
            <div class="kv"><b>Natureza</b><span>${{safe(row.natureza_juridica)}}</span></div>
          </div>
          <div class="panel"><h3>Fiscal e credito</h3>
            <div class="kv"><b>Regime</b><span>${{row.regime_label}}</span></div>
            <div class="kv"><b>Credito</b><span>${{row.credito_label}}</span></div>
            <div class="kv"><b>Simples</b><span>${{safe(row.opcao_simples)}}</span></div>
            <div class="kv"><b>MEI</b><span>${{safe(row.opcao_mei)}}</span></div>
            <div class="kv"><b>Validacao</b><span>${{row.validacao_label}}</span></div>
            <div class="tags">${{fiscalTags(row)}}</div>
          </div>
          <div class="panel"><h3>Operacao e saneamento</h3>
            <div class="kv"><b>Curva</b><span>${{row.curva_label}} | pos. ${{safe(row.curva_posicao)}} | ${{row.curva_valor_fmt}}</span></div>
            <div class="kv"><b>NFe</b><span>${{safe(row.nfe_linhas)}} linhas | ${{row.nfe_valor_fmt}}</span></div>
            <div class="kv"><b>UFs</b><span>${{safe(row.nfe_ufs)}}</span></div>
            <div class="kv"><b>Anos</b><span>${{safe(row.nfe_anos)}}</span></div>
            <div class="kv"><b>Saneamento</b><span>${{row.saneamento_label}}</span></div>
            <div class="tags">${{operacaoTags(row)}}</div>
          </div>
        </div>
      </td>
    </tr>`;
}}

function renderStats(rows) {{
  els.stRows.textContent = rows.length.toLocaleString('pt-BR');
  els.stNew.textContent = rows.filter(r => r.dado_novo_label === 'Com dado novo').length.toLocaleString('pt-BR');
  els.stMov.textContent = rows.filter(r => r.movimento_label === 'Com movimento').length.toLocaleString('pt-BR');
  els.stReview.textContent = rows.filter(r => r.acao_label === 'Revisar antes').length.toLocaleString('pt-BR');
  els.stFiscal.textContent = rows.filter(r => r.acao_label === 'Validar fiscal').length.toLocaleString('pt-BR');
}}

function totalPages(rows) {{
  return Math.max(1, Math.ceil(rows.length / state.pageSize));
}}

function updatePagination(rows) {{
  const total = totalPages(rows);
  if (state.page > total) state.page = total;
  if (state.page < 1) state.page = 1;
  els.pageInput.value = state.page;
  els.pageInput.max = total;
  els.firstPage.disabled = state.page === 1;
  els.prevPage.disabled = state.page === 1;
  els.nextPage.disabled = state.page === total;
  els.lastPage.disabled = state.page === total;
}}

function render() {{
  const rows = filteredRows();
  renderStats(rows);
  updatePagination(rows);
  const start = (state.page - 1) * state.pageSize;
  const pageRows = rows.slice(start, start + state.pageSize);
  els.empty.style.display = rows.length ? 'none' : 'block';
  els.tbody.innerHTML = pageRows.map(summaryRow).join('');
  els.tbody.querySelectorAll('tr.summary').forEach(tr => {{
    tr.addEventListener('click', () => {{
      const key = tr.dataset.key;
      const wasOpen = state.expandedKey === key;
      state.expandedKey = wasOpen ? '' : key;
      els.tbody.querySelectorAll('tr.detail-row.open').forEach(row => row.classList.remove('open'));
      const detail = els.tbody.querySelector(`tr[data-detail="${{key}}"]`);
      if (detail && !wasOpen) detail.classList.add('open');
    }});
  }});
  applyColumnVisibility();
}}

function applyColumnVisibility() {{
  const visibleLabels = selected('campos');
  columns.forEach(([key, label]) => {{
    document.querySelectorAll(`[data-col="${{key}}"]`).forEach(el => {{
      el.style.display = (!visibleLabels.length || visibleLabels.includes(label)) ? '' : 'none';
    }});
  }});
}}

function openModal(id) {{
  document.getElementById(id).classList.add('open');
}}

function closeModal(id) {{
  document.getElementById(id).classList.remove('open');
}}

function updateStickyOffset() {{
  const panel = document.getElementById('stickyPanel');
  const height = panel ? Math.ceil(panel.getBoundingClientRect().height) : 0;
  document.documentElement.style.setProperty('--sticky-offset', `${{height}}px`);
}}

function installHelpText() {{
  const filterHelp = {{
    q: 'Pesquisa textual livre sobre a base consolidada. Procura em codigo interno, CNPJ/CPF sem formatacao, CNPJ/CPF formatado, razao social original, fantasia original, emails internos, razao oficial, fantasia oficial, email oficial e acao sugerida. Use quando voce ja sabe parte do nome, documento ou email e quer localizar rapidamente o fornecedor.',
    empresa: 'Filtra a empresa do grupo em que o fornecedor aparece no cadastro original. Um mesmo fornecedor pode aparecer em mais de uma empresa; nesse caso, ele permanece em uma unica linha consolidada e recebe tags coloridas: IDEAL azul, MELHOR verde e POMME vermelho.',
    tipo: 'Classifica o fornecedor a partir do prefixo do codigo interno. Codigos iniciados por J sao tratados como PJ/CNPJ, codigos iniciados por F como PF/CPF, e qualquer outro prefixo entra como Outro para revisao operacional.',
    movimento: 'Separa fornecedores que tiveram notas fiscais de entrada encontradas na base NFE dos fornecedores que existem apenas no cadastro. Com movimento indica historico de compras; sem movimento indica cadastro sem nota localizada no periodo analisado.',
    curva: 'Classificacao de relevancia economica por valor comprado na curva de fornecedores. AAA representa fornecedores de maior peso financeiro, seguido de AA, A, B, BB, C, CC e CCC. Sem curva significa que o fornecedor nao apareceu na base de curva consolidada.',
    situacao: 'Situacao cadastral externa obtida ou inferida a partir da consulta OpenCNPJ. Ativa indica cadastro oficial ativo; Nao ativa agrega situacoes como baixada, inapta, suspensa ou similares; Nao consultado indica que nao houve consulta valida; Sem dado indica ausencia de campo suficiente para classificacao.',
    regime: 'Leitura fiscal inicial do fornecedor. Simples indica opcao pelo Simples Nacional; Nao Simples indica fornecedor fora do Simples; MEI indica microempreendedor individual; Indeterminado significa que a consulta nao permitiu concluir o regime; Sem dado significa que a fase ainda nao trouxe informacao fiscal util.',
    credito: 'Estimativa inicial de potencial de credito na logica fiscal de 2027. Alto indica maior chance de gerar credito relevante; Condicionado indica dependencia de regra ou opcao fiscal, especialmente no Simples; Baixo indica potencial reduzido ou inexistente; Indeterminado exige validacao; Sem dado nao recebeu analise fiscal nesta fase.',
    dadoNovo: 'Indica se a etapa de enriquecimento acrescentou informacoes que nao estavam no cadastro original ou que estavam vazias. Com dado novo significa que ha algum campo oficial ou analitico novo disponivel para eventual retorno ao BD; Sem dado novo significa que nada util foi acrescentado nesta etapa.',
    divergencia: 'Compara campos cadastrais internos com campos oficiais retornados. Razao social, fantasia e email filtram divergencias especificas; Qualquer divergencia captura qualquer uma dessas diferencas; Sem divergencia mostra fornecedores sem conflito cadastral nesses campos comparados.',
    saneamento: 'Mostra a necessidade de completar ou corrigir o cadastro interno. Precisa sanear indica falta de informacoes essenciais, como email ou inscricoes quando aplicavel; Nao precisa indica ausencia de pendencia essencial detectada; Sem analise indica que nao houve insumo suficiente para concluir.',
    validacao: 'Aponta fornecedores que precisam de revisao fiscal antes de uso automatico. Precisa validar inclui regime indeterminado, conflito entre cadastro interno e oficial ou outra incerteza fiscal; Nao precisa indica que a leitura inicial nao levantou alerta; Sem analise indica falta de base para avaliar.',
    acao: 'Proxima acao operacional sugerida. Atualizar BD significa dado novo sem alerta relevante; Revisar antes significa que ha divergencia a conferir; Validar fiscal prioriza incerteza tributaria; Sanear cadastro prioriza campos faltantes; Sem alteracao indica que a fase nao gerou mudanca acionavel.',
    completude: 'Faixa do score cadastral calculado na fase de qualidade cadastral. Alta indica cadastro relativamente completo; media indica cadastro usavel, mas com lacunas; baixa indica falta relevante de campos essenciais ou complementares. O score mede qualidade cadastral, nao regularidade fiscal definitiva.',
    prioridade: 'Prioridade de tratamento combinando importancia economica e risco operacional. A faixa considera movimento em NFe, posicao na curva, valor comprado, pendencias cadastrais, divergencias e necessidade fiscal. Critica e alta devem ser tratadas antes de media e baixa.',
    campos: 'Controla quais colunas resumidas aparecem na tabela principal. Use para simplificar a leitura quando estiver focando em fiscal, credito, movimento ou acao. A selecao altera apenas a visualizacao; nao remove dados do painel nem dos detalhes.'
  }};
  Object.entries(filterHelp).forEach(([id, tip]) => {{
    const help = document.getElementById(id)?.closest('.filter')?.querySelector('.help');
    if (help) help.dataset.tip = tip;
  }});
  const columnHelp = {{
    fornecedor: 'Fornecedor consolidado por documento/chave. Mostra o nome de exibicao, as empresas do grupo onde o fornecedor aparece e o CNPJ/CPF formatado. A linha pode representar mais de uma linha original do cadastro.',
    mudanca: 'Resumo do enriquecimento e das divergencias. Indica se foram incluidos dados oficiais ou analiticos e se houve conflito entre o cadastro interno e a fonte externa.',
    fiscal: 'Situacao oficial e regime fiscal indicado. Serve como leitura inicial de risco, nao como parecer fiscal definitivo.',
    credito: 'Potencial inicial de credito considerado para a logica de 2027. A classificacao depende do regime, da qualidade do dado externo e da necessidade de validacao fiscal.',
    movimento: 'Mostra a relevancia economica: curva, valor da curva e valor de NFe encontrado. Ajuda a priorizar fornecedores que afetam mais compras e operacao.',
    acao: 'Acao sugerida para a proxima etapa: atualizar BD, revisar antes, validar fiscal, sanear cadastro ou manter sem alteracao.',
    prioridade: 'Score cadastral, faixa de completude e prioridade de saneamento. Combina qualidade do cadastro com importancia operacional.'
  }};
  Object.entries(columnHelp).forEach(([key, tip]) => {{
    const help = document.querySelector(`th[data-col="${{key}}"] .help`);
    if (help) help.dataset.tip = tip;
  }});
}}

installHelpText();
initFilters();
updateStickyOffset();
window.addEventListener('resize', updateStickyOffset);
els.q.addEventListener('input', () => {{
  els.q.closest('.filter')?.classList.toggle('active', Boolean(els.q.value.trim()));
  state.page = 1;
  state.expandedKey = '';
  render();
}});
els.pageSize.addEventListener('change', () => {{
  state.pageSize = Number(els.pageSize.value);
  state.page = 1;
  state.expandedKey = '';
  render();
}});
els.pageInput.addEventListener('change', () => {{
  state.page = Number(els.pageInput.value || 1);
  state.expandedKey = '';
  render();
}});
els.firstPage.addEventListener('click', () => {{
  state.page = 1;
  state.expandedKey = '';
  render();
}});
els.prevPage.addEventListener('click', () => {{
  state.page -= 1;
  state.expandedKey = '';
  render();
}});
els.nextPage.addEventListener('click', () => {{
  state.page += 1;
  state.expandedKey = '';
  render();
}});
els.lastPage.addEventListener('click', () => {{
  state.page = totalPages(filteredRows());
  state.expandedKey = '';
  render();
}});
document.addEventListener('click', () => document.querySelectorAll('.multi.open').forEach(open => open.classList.remove('open')));
document.getElementById('methodBtn').addEventListener('click', () => openModal('methodModal'));
document.querySelectorAll('[data-close]').forEach(btn => btn.addEventListener('click', () => closeModal(btn.dataset.close)));
document.querySelectorAll('.modal-backdrop').forEach(backdrop => {{
  backdrop.addEventListener('click', event => {{
    if (event.target === backdrop) backdrop.classList.remove('open');
  }});
}});
render();
</script>
</body>
</html>"""
    AUDIT_HTML_FILE.write_text(html, encoding="utf-8")


def main() -> None:
    PHASE_04_DIR.mkdir(parents=True, exist_ok=True)
    rows = build_rows()
    build_html(rows)
    print(f"Arquivo gerado: {AUDIT_HTML_FILE}")
    print(f"Registros: {len(rows)}")


if __name__ == "__main__":
    main()
