from __future__ import annotations

import json
import re
from copy import copy
from html import escape
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
PHASE_01_DIR = BASE_DIR / "output" / "fase_01_cadastro"
PHASE_02_DIR = BASE_DIR / "output" / "fase_02_reconciliacao"
PHASE_03_DIR = BASE_DIR / "output" / "fase_03_enriquecimento"
PHASE_04_DIR = BASE_DIR / "output" / "fase_04_visualizacao"

MASTER_FILE = DATA_DIR / "FORNECEDORES - TODAS AS EMPRESAS.xlsx"
NORMALIZED_FILE = PHASE_01_DIR / "01_fornecedores_cadastro_normalizado.csv"
RECONCILED_FILE = PHASE_02_DIR / "01_fornecedores_reconciliados_movimento.csv"
CLASSIFIED_FILE = PHASE_03_DIR / "03_fornecedores_classificacao_fiscal_inicial_lote_completo.csv"
ENRICHED_RAW_FILE = PHASE_03_DIR / "01_fornecedores_opencnpj_enriquecidos_lote_completo.csv"

COMPARE_WORKBOOK = PHASE_04_DIR / "01_antes_depois_visual_fornecedores.xlsx"
RETURN_WORKBOOK = PHASE_04_DIR / "02_fornecedores_formato_original_com_campos_gvg.xlsx"
HTML_FILE = PHASE_04_DIR / "03_painel_interativo_fornecedores.html"

COMPARE_HEADERS = [
    "LINHA_ORIGINAL",
    "EMPRESA",
    "CODIGO",
    "TIPO_CADASTRO",
    "DOCUMENTO",
    "ANTES_RAZAO_SOCIAL",
    "DEPOIS_RAZAO_SOCIAL_OFICIAL",
    "ANTES_NOME_FANTASIA",
    "DEPOIS_NOME_FANTASIA_OFICIAL",
    "ANTES_STATUS_INTERNO",
    "DEPOIS_STATUS_OFICIAL",
    "ANTES_EMAILS",
    "DEPOIS_EMAIL_OFICIAL",
    "MOVIMENTO_NFE",
    "VALOR_NFE_TOTAL",
    "CURVA_FORN",
    "SCORE_CADASTRAL",
    "FAIXA_CADASTRAL",
    "REGIME_INDICIADO",
    "POTENCIAL_CREDITO_2027",
    "SANEAMENTO_CADASTRAL",
    "VALIDACAO_FISCAL",
    "MOTIVOS_SANEAMENTO",
    "MOTIVOS_VALIDACAO",
]

RETURN_HEADERS = [
    "GVG_TIPO_CADASTRO",
    "GVG_CHAVE_UNIFICACAO",
    "GVG_SCORE_CADASTRAL",
    "GVG_FAIXA_CADASTRAL",
    "GVG_PENDENCIAS_ESSENCIAIS",
    "GVG_PENDENCIAS_COMPLEMENTARES",
    "GVG_TEM_MOVIMENTO_NFE",
    "GVG_NFE_LINHAS",
    "GVG_NFE_VALOR_TOTAL",
    "GVG_CURVA_CLASSIFICACAO",
    "GVG_PRIORIDADE_SANEAMENTO_SCORE",
    "GVG_PRIORIDADE_SANEAMENTO_FAIXA",
    "GVG_SITUACAO_CADASTRAL_OFICIAL",
    "GVG_MATRIZ_FILIAL",
    "GVG_PORTE_EMPRESA",
    "GVG_NATUREZA_JURIDICA",
    "GVG_REGIME_INDICIADO",
    "GVG_POTENCIAL_CREDITO_2027",
    "GVG_NECESSITA_SANEAMENTO",
    "GVG_MOTIVOS_SANEAMENTO",
    "GVG_NECESSITA_VALIDACAO_FISCAL",
    "GVG_MOTIVOS_VALIDACAO_FISCAL",
    "GVG_RAZAO_SOCIAL_OFICIAL",
    "GVG_NOME_FANTASIA_OFICIAL",
    "GVG_EMAIL_OFICIAL",
]

FILL_HEADER_BEFORE = PatternFill("solid", fgColor="FCE4D6")
FILL_HEADER_AFTER = PatternFill("solid", fgColor="D9EAF7")
FILL_HEADER_ANALYSIS = PatternFill("solid", fgColor="E2F0D9")
FILL_HEADER_RISK = PatternFill("solid", fgColor="FFF2CC")
FILL_CHANGED = PatternFill("solid", fgColor="FFE699")
FILL_ADDED = PatternFill("solid", fgColor="D9EAF7")
FILL_OK = PatternFill("solid", fgColor="C6E0B4")
FILL_WARN = PatternFill("solid", fgColor="FFD966")
FILL_RISK = PatternFill("solid", fgColor="F4CCCC")
FILL_NEUTRAL = PatternFill("solid", fgColor="EDEDED")
FILL_GVG = PatternFill("solid", fgColor="DDEBF7")
FILL_GVG_HEADER = PatternFill("solid", fgColor="5B9BD5")
FONT_WHITE = Font(color="FFFFFF", bold=True)
FONT_BOLD = Font(bold=True)
ALIGN_TOP = Alignment(vertical="top", wrap_text=True)


def normalize_cnpj(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits if len(digits) == 14 else ""


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_text_upper(value: Any) -> str:
    return normalize_text(value).upper()


def classify_code_type(code: str) -> str:
    prefix = str(code or "").strip()[:1].upper()
    if prefix == "J":
        return "PJ/CNPJ"
    if prefix == "F":
        return "PF/CPF"
    if prefix == "I":
        return "INSCRICAO/OUTRO"
    if prefix == "O":
        return "OUTRO"
    return "NAO_IDENTIFICADO"


def build_key(sheet_name: str, codigo: str, cnpj: str, razao: str) -> tuple[str, str, str, str]:
    return (sheet_name, codigo, cnpj, razao)


def load_csv_index(path: Path, key_fields: list[str]) -> dict[tuple[str, ...], dict[str, str]]:
    import csv

    out: dict[tuple[str, ...], dict[str, str]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            out[tuple(row[field] for field in key_fields)] = row
    return out


def load_csv_by_field(path: Path, field: str) -> dict[str, dict[str, str]]:
    import csv

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return {row[field]: row for row in csv.DictReader(handle)}


def map_fill_for_credit(value: str) -> PatternFill:
    text = normalize_text(value)
    if text == "alto_potencial":
        return FILL_OK
    if text == "potencial_condicionado":
        return FILL_WARN
    if text == "baixo_ou_inexistente":
        return FILL_RISK
    return FILL_NEUTRAL


def map_fill_for_yes_no(value: str) -> PatternFill:
    return FILL_WARN if str(value).strip().upper() == "S" else FILL_OK


def set_widths(ws) -> None:
    widths = {
        "A": 11,
        "B": 12,
        "C": 18,
        "D": 14,
        "E": 18,
        "F": 34,
        "G": 38,
        "H": 26,
        "I": 30,
        "J": 18,
        "K": 18,
        "L": 26,
        "M": 26,
        "N": 12,
        "O": 16,
        "P": 12,
        "Q": 14,
        "R": 14,
        "S": 18,
        "T": 20,
        "U": 18,
        "V": 18,
        "W": 34,
        "X": 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_compare_workbook(rows_by_sheet: dict[str, list[dict[str, Any]]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    legend = wb.create_sheet("LEGENDA", 0)
    legend["A1"] = "Planilha visual de antes x depois"
    legend["A1"].font = FONT_BOLD
    legend["A3"] = "Laranja"
    legend["B3"] = "Campo original"
    legend["A4"] = "Azul"
    legend["B4"] = "Campo oficial/enriquecido"
    legend["A5"] = "Amarelo"
    legend["B5"] = "Divergencia ou atencao"
    legend["A6"] = "Verde"
    legend["B6"] = "Situacao favoravel/confirmada"
    legend["A7"] = "Vermelho"
    legend["B7"] = "Risco mais alto"
    legend["A8"] = "Cinza"
    legend["B8"] = "Sem informacao suficiente"
    legend["A3"].fill = FILL_HEADER_BEFORE
    legend["A4"].fill = FILL_HEADER_AFTER
    legend["A5"].fill = FILL_WARN
    legend["A6"].fill = FILL_OK
    legend["A7"].fill = FILL_RISK
    legend["A8"].fill = FILL_NEUTRAL
    legend.column_dimensions["A"].width = 16
    legend.column_dimensions["B"].width = 48

    for sheet_name, rows in rows_by_sheet.items():
        ws = wb.create_sheet(sheet_name)
        for col_idx, header in enumerate(COMPARE_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = FONT_BOLD
            cell.alignment = ALIGN_TOP
            if "ANTES_" in header:
                cell.fill = FILL_HEADER_BEFORE
            elif "DEPOIS_" in header:
                cell.fill = FILL_HEADER_AFTER
            elif header in {"MOVIMENTO_NFE", "VALOR_NFE_TOTAL", "CURVA_FORN", "SCORE_CADASTRAL", "FAIXA_CADASTRAL", "REGIME_INDICIADO", "POTENCIAL_CREDITO_2027"}:
                cell.fill = FILL_HEADER_ANALYSIS
            else:
                cell.fill = FILL_HEADER_RISK

        for row_idx, item in enumerate(rows, start=2):
            values = [
                item["linha_original"],
                item["empresa_origem"],
                item["codigo_fornecedor"],
                item["tipo_cadastro"],
                item["documento"],
                item["antes_razao_social"],
                item["depois_razao_social_oficial"],
                item["antes_nome_fantasia"],
                item["depois_nome_fantasia_oficial"],
                item["antes_status_interno"],
                item["depois_status_oficial"],
                item["antes_emails"],
                item["depois_email_oficial"],
                item["movimento_nfe"],
                item["nfe_valor_total"],
                item["curva_forn"],
                item["score_cadastral"],
                item["faixa_cadastral"],
                item["regime_indiciado"],
                item["potencial_credito_2027"],
                item["necessita_saneamento_cadastral"],
                item["necessita_validacao_fiscal"],
                item["motivos_saneamento_cadastral"],
                item["motivos_validacao_fiscal"],
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = ALIGN_TOP

            if item["divergencia_razao_social"] == "S":
                ws.cell(row_idx, 6).fill = FILL_CHANGED
                ws.cell(row_idx, 7).fill = FILL_CHANGED
            else:
                ws.cell(row_idx, 6).fill = FILL_HEADER_BEFORE
                ws.cell(row_idx, 7).fill = FILL_HEADER_AFTER

            if item["divergencia_nome_fantasia"] == "S":
                ws.cell(row_idx, 8).fill = FILL_CHANGED
                ws.cell(row_idx, 9).fill = FILL_CHANGED
            else:
                ws.cell(row_idx, 8).fill = FILL_HEADER_BEFORE
                ws.cell(row_idx, 9).fill = FILL_HEADER_AFTER

            ws.cell(row_idx, 10).fill = FILL_HEADER_BEFORE
            status_fill = FILL_OK if normalize_text(item["depois_status_oficial"]).lower() == "ativa" else (FILL_RISK if item["depois_status_oficial"] else FILL_NEUTRAL)
            ws.cell(row_idx, 11).fill = status_fill
            if item["divergencia_email"] == "S":
                ws.cell(row_idx, 12).fill = FILL_CHANGED
                ws.cell(row_idx, 13).fill = FILL_CHANGED
            else:
                ws.cell(row_idx, 12).fill = FILL_HEADER_BEFORE
                ws.cell(row_idx, 13).fill = FILL_HEADER_AFTER

            ws.cell(row_idx, 14).fill = FILL_OK if item["movimento_nfe"] == "S" else FILL_NEUTRAL
            ws.cell(row_idx, 15).fill = FILL_OK if float(item["nfe_valor_total"] or 0) > 0 else FILL_NEUTRAL
            ws.cell(row_idx, 16).fill = FILL_OK if item["curva_forn"] else FILL_NEUTRAL
            ws.cell(row_idx, 17).fill = FILL_HEADER_ANALYSIS
            ws.cell(row_idx, 18).fill = FILL_HEADER_ANALYSIS
            ws.cell(row_idx, 19).fill = map_fill_for_credit(item["regime_indiciado"])
            ws.cell(row_idx, 20).fill = map_fill_for_credit(item["potencial_credito_2027"])
            ws.cell(row_idx, 21).fill = map_fill_for_yes_no(item["necessita_saneamento_cadastral"])
            ws.cell(row_idx, 22).fill = map_fill_for_yes_no(item["necessita_validacao_fiscal"])
            ws.cell(row_idx, 23).fill = FILL_HEADER_RISK
            ws.cell(row_idx, 24).fill = FILL_HEADER_RISK

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        set_widths(ws)

    wb.save(COMPARE_WORKBOOK)


def build_return_workbook(source_rows_by_sheet: dict[str, list[dict[str, Any]]]) -> None:
    wb = load_workbook(MASTER_FILE)
    for ws in wb.worksheets:
        rows = source_rows_by_sheet.get(ws.title, [])
        if not rows:
            continue
        start_col = ws.max_column + 1
        for offset, header in enumerate(RETURN_HEADERS):
            cell = ws.cell(row=1, column=start_col + offset, value=header)
            cell.fill = FILL_GVG_HEADER
            cell.font = FONT_WHITE
            cell.alignment = ALIGN_TOP

        for item in rows:
            row_idx = item["linha_original"]
            values = [
                item["tipo_cadastro"],
                item["chave_unificacao"],
                item["score_cadastral"],
                item["faixa_cadastral"],
                item["pendencias_essenciais"],
                item["pendencias_complementares"],
                item["movimento_nfe"],
                item["nfe_linhas"],
                item["nfe_valor_total"],
                item["curva_forn"],
                item["prioridade_saneamento_score"],
                item["prioridade_saneamento_faixa"],
                item["depois_status_oficial"],
                item["matriz_filial"],
                item["porte_empresa"],
                item["natureza_juridica"],
                item["regime_indiciado"],
                item["potencial_credito_2027"],
                item["necessita_saneamento_cadastral"],
                item["motivos_saneamento_cadastral"],
                item["necessita_validacao_fiscal"],
                item["motivos_validacao_fiscal"],
                item["depois_razao_social_oficial"],
                item["depois_nome_fantasia_oficial"],
                item["depois_email_oficial"],
            ]
            for offset, value in enumerate(values):
                cell = ws.cell(row=row_idx, column=start_col + offset, value=value)
                cell.alignment = ALIGN_TOP
                cell.fill = FILL_GVG

            ws.cell(row_idx, start_col + 12).fill = FILL_OK if normalize_text(item["depois_status_oficial"]).lower() == "ativa" else (FILL_RISK if item["depois_status_oficial"] else FILL_NEUTRAL)
            ws.cell(row_idx, start_col + 16).fill = map_fill_for_credit(item["regime_indiciado"])
            ws.cell(row_idx, start_col + 17).fill = map_fill_for_credit(item["potencial_credito_2027"])
            ws.cell(row_idx, start_col + 18).fill = map_fill_for_yes_no(item["necessita_saneamento_cadastral"])
            ws.cell(row_idx, start_col + 20).fill = map_fill_for_yes_no(item["necessita_validacao_fiscal"])

        for offset in range(len(RETURN_HEADERS)):
            ws.column_dimensions[chr(64 + start_col + offset) if start_col + offset <= 26 else ws.cell(1, start_col + offset).column_letter].width = 18
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(RETURN_WORKBOOK)


def build_html(rows: list[dict[str, Any]]) -> None:
    payload = json.dumps(rows, ensure_ascii=False)
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<title>Painel Interativo de Fornecedores</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
body{{font-family:Segoe UI,Arial,sans-serif;margin:0;background:#f4f7fb;color:#1f2937}}
.app{{min-height:100vh}}
.main{{padding:18px}}
.page-title{{margin:0 0 18px;font-size:32px;font-weight:800;color:#0f172a}}
.filters{{display:grid;grid-template-columns:repeat(4,minmax(160px,1fr));gap:12px;margin-bottom:16px}}
.filters input,.filters select{{width:100%;padding:10px;border:1px solid #cbd5e1;border-radius:10px;background:#fff}}
.cards{{display:grid;grid-template-columns:repeat(4,minmax(160px,1fr));gap:12px;margin-bottom:16px}}
.card{{background:#fff;border-radius:14px;padding:14px;box-shadow:0 8px 24px rgba(15,23,42,.06)}}
.card .label{{font-size:12px;color:#64748b;text-transform:uppercase}}
.card .value{{font-size:24px;font-weight:700;margin-top:6px}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:14px;overflow:hidden;box-shadow:0 8px 24px rgba(15,23,42,.06)}}
th,td{{padding:10px 12px;border-bottom:1px solid #e5e7eb;text-align:left;font-size:13px;vertical-align:top}}
th{{background:#eaf1fb;color:#0f172a;position:sticky;top:0}}
tr:hover{{background:#f8fbff}}
.pill{{display:inline-block;padding:4px 8px;border-radius:999px;font-size:12px;font-weight:600}}
.high{{background:#dcfce7;color:#166534}}
.mid{{background:#fef3c7;color:#92400e}}
.risk{{background:#fee2e2;color:#991b1b}}
.neutral{{background:#e5e7eb;color:#334155}}
.detail{{margin-top:16px;background:#fff;border-radius:14px;padding:18px;box-shadow:0 8px 24px rgba(15,23,42,.06)}}
.detail-grid{{display:grid;grid-template-columns:repeat(2,minmax(220px,1fr));gap:16px}}
.box{{border:1px solid #e5e7eb;border-radius:12px;padding:12px}}
.box h3{{margin:0 0 8px;font-size:14px}}
.mono{{font-family:Consolas,monospace}}
</style>
</head>
<body>
<div class="app">
  <main class="main">
    <h1 class="page-title">Painel de Fornecedores</h1>
    <div class="filters">
      <input id="q" placeholder="Buscar por codigo, CNPJ, nome">
      <select id="empresa"></select>
      <select id="movimento"></select>
      <select id="credito"></select>
    </div>
    <div class="cards">
      <div class="card"><div class="label">Linhas</div><div class="value" id="countRows">0</div></div>
      <div class="card"><div class="label">Com Movimento</div><div class="value" id="countMov">0</div></div>
      <div class="card"><div class="label">Alto Potencial</div><div class="value" id="countHigh">0</div></div>
      <div class="card"><div class="label">Validacao Fiscal</div><div class="value" id="countVal">0</div></div>
    </div>
    <table>
      <thead>
        <tr>
          <th>Empresa</th>
          <th>Codigo</th>
          <th>Documento</th>
          <th>Antes</th>
          <th>Depois</th>
          <th>Mov.</th>
          <th>Curva</th>
          <th>Regime</th>
          <th>Credito</th>
          <th>Saneamento</th>
          <th>Validacao</th>
        </tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
    <section class="detail" id="detail">
      <strong>Selecione uma linha para ver o detalhe.</strong>
    </section>
  </main>
</div>
<script>
const DATA = {payload};
const els = {{
  q: document.getElementById('q'),
  empresa: document.getElementById('empresa'),
  movimento: document.getElementById('movimento'),
  credito: document.getElementById('credito'),
  tbody: document.getElementById('tbody'),
  detail: document.getElementById('detail'),
  countRows: document.getElementById('countRows'),
  countMov: document.getElementById('countMov'),
  countHigh: document.getElementById('countHigh'),
  countVal: document.getElementById('countVal'),
}};

function unique(values) {{
  return [...new Set(values.filter(Boolean))].sort();
}}

function fillSelect(el, values, label) {{
  el.innerHTML = `<option value="">${{label}}</option>` + values.map(v => `<option value="${{v}}">${{v}}</option>`).join('');
}}

function badgeClass(value) {{
  if (value === 'alto_potencial') return 'high';
  if (value === 'potencial_condicionado') return 'mid';
  if (value === 'baixo_ou_inexistente') return 'risk';
  return 'neutral';
}}

function ynClass(value) {{
  return value === 'S' ? 'risk' : 'high';
}}

fillSelect(els.empresa, unique(DATA.map(r => r.empresa_origem)), 'Todas as empresas');
fillSelect(els.movimento, ['', 'S', 'N'], 'Movimento NFE');
fillSelect(els.credito, unique(DATA.map(r => r.potencial_credito_2027)), 'Todos os creditos');

function filtered() {{
  const q = els.q.value.trim().toLowerCase();
  return DATA.filter(row => {{
    if (els.empresa.value && row.empresa_origem !== els.empresa.value) return false;
    if (els.movimento.value && row.movimento_nfe !== els.movimento.value) return false;
    if (els.credito.value && row.potencial_credito_2027 !== els.credito.value) return false;
    if (!q) return true;
    const hay = [
      row.codigo_fornecedor, row.documento, row.antes_razao_social, row.depois_razao_social_oficial,
      row.antes_nome_fantasia, row.depois_nome_fantasia_oficial
    ].join(' ').toLowerCase();
    return hay.includes(q);
  }});
}}

function renderDetail(row) {{
  els.detail.innerHTML = `
    <div class="detail-grid">
      <div class="box">
        <h3>Antes</h3>
        <div><strong>Empresa:</strong> ${{row.empresa_origem}}</div>
        <div><strong>Codigo:</strong> <span class="mono">${{row.codigo_fornecedor}}</span></div>
        <div><strong>Documento:</strong> <span class="mono">${{row.documento}}</span></div>
        <div><strong>Razao social:</strong> ${{row.antes_razao_social || '—'}}</div>
        <div><strong>Nome fantasia:</strong> ${{row.antes_nome_fantasia || '—'}}</div>
        <div><strong>Status interno:</strong> ${{row.antes_status_interno || '—'}}</div>
        <div><strong>Emails:</strong> ${{row.antes_emails || '—'}}</div>
      </div>
      <div class="box">
        <h3>Depois</h3>
        <div><strong>Razao oficial:</strong> ${{row.depois_razao_social_oficial || '—'}}</div>
        <div><strong>Fantasia oficial:</strong> ${{row.depois_nome_fantasia_oficial || '—'}}</div>
        <div><strong>Status oficial:</strong> ${{row.depois_status_oficial || '—'}}</div>
        <div><strong>Email oficial:</strong> ${{row.depois_email_oficial || '—'}}</div>
        <div><strong>Regime:</strong> ${{row.regime_indiciado || '—'}}</div>
        <div><strong>Credito 2027:</strong> ${{row.potencial_credito_2027 || '—'}}</div>
      </div>
      <div class="box">
        <h3>Operacao</h3>
        <div><strong>Movimento NFE:</strong> ${{row.movimento_nfe}}</div>
        <div><strong>Linhas NFE:</strong> ${{row.nfe_linhas}}</div>
        <div><strong>Valor NFE:</strong> ${{row.nfe_valor_total}}</div>
        <div><strong>Curva:</strong> ${{row.curva_forn || '—'}}</div>
        <div><strong>Score cadastral:</strong> ${{row.score_cadastral}}</div>
      </div>
      <div class="box">
        <h3>Risco</h3>
        <div><strong>Saneamento cadastral:</strong> ${{row.necessita_saneamento_cadastral}}</div>
        <div><strong>Motivos saneamento:</strong> ${{row.motivos_saneamento_cadastral || '—'}}</div>
        <div><strong>Validacao fiscal:</strong> ${{row.necessita_validacao_fiscal}}</div>
        <div><strong>Motivos validacao:</strong> ${{row.motivos_validacao_fiscal || '—'}}</div>
      </div>
    </div>`;
}}

function render() {{
  const rows = filtered();
  els.countRows.textContent = rows.length;
  els.countMov.textContent = rows.filter(r => r.movimento_nfe === 'S').length;
  els.countHigh.textContent = rows.filter(r => r.potencial_credito_2027 === 'alto_potencial').length;
  els.countVal.textContent = rows.filter(r => r.necessita_validacao_fiscal === 'S').length;
  els.tbody.innerHTML = rows.slice(0, 1000).map((row, index) => `
    <tr data-index="${{index}}">
      <td>${{row.empresa_origem}}</td>
      <td class="mono">${{row.codigo_fornecedor}}</td>
      <td class="mono">${{row.documento}}</td>
      <td>${{row.antes_razao_social || '—'}}</td>
      <td>${{row.depois_razao_social_oficial || '—'}}</td>
      <td><span class="pill ${{row.movimento_nfe === 'S' ? 'high' : 'neutral'}}">${{row.movimento_nfe}}</span></td>
      <td>${{row.curva_forn || '—'}}</td>
      <td>${{row.regime_indiciado || '—'}}</td>
      <td><span class="pill ${{badgeClass(row.potencial_credito_2027)}}">${{row.potencial_credito_2027 || '—'}}</span></td>
      <td><span class="pill ${{ynClass(row.necessita_saneamento_cadastral)}}">${{row.necessita_saneamento_cadastral}}</span></td>
      <td><span class="pill ${{ynClass(row.necessita_validacao_fiscal)}}">${{row.necessita_validacao_fiscal}}</span></td>
    </tr>`).join('');
  [...els.tbody.querySelectorAll('tr')].forEach((tr, idx) => tr.addEventListener('click', () => renderDetail(rows[idx])));
  if (rows.length) renderDetail(rows[0]);
}}

[els.q, els.empresa, els.movimento, els.credito].forEach(el => el.addEventListener('input', render));
render();
</script>
</body>
</html>"""
    HTML_FILE.write_text(html, encoding="utf-8")


def build_html_v2(rows: list[dict[str, Any]]) -> None:
    payload = json.dumps(rows, ensure_ascii=False)
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<title>Painel Interativo de Fornecedores</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
*{{box-sizing:border-box}}
body{{font-family:Segoe UI,Arial,sans-serif;margin:0;background:#eef3f9;color:#1f2937}}
.app{{min-height:100vh}}
.main{{padding:20px}}
.page-title{{margin:0 0 16px;font-size:32px;font-weight:800;color:#0f172a}}
.filters{{display:grid;grid-template-columns:repeat(4,minmax(180px,1fr));gap:12px;margin-bottom:16px}}
.filters input,.filters select{{width:100%;padding:12px;border:1px solid #cbd5e1;border-radius:12px;background:#fff;font-size:14px}}
.cards{{display:grid;grid-template-columns:repeat(4,minmax(180px,1fr));gap:12px;margin-bottom:16px}}
.card{{background:#fff;border-radius:16px;padding:16px;box-shadow:0 10px 30px rgba(15,23,42,.07)}}
.card .label{{font-size:12px;color:#64748b;text-transform:uppercase}}
.card .value{{font-size:24px;font-weight:700;margin-top:6px}}
.right-note{{font-size:12px;color:#64748b;margin-bottom:10px}}
.table-wrap{{background:#fff;border-radius:18px;overflow:auto;box-shadow:0 10px 30px rgba(15,23,42,.07)}}
table{{width:100%;border-collapse:separate;border-spacing:0;min-width:1700px}}
th,td{{padding:12px 14px;border-bottom:1px solid #e5e7eb;text-align:left;font-size:13px;vertical-align:top}}
th{{position:sticky;top:0;z-index:2;color:#0f172a;font-weight:700}}
th.meta{{background:#dbe7f5}}
th.before{{background:#fce4d6}}
th.after{{background:#d9eaf7}}
th.analysis{{background:#e2f0d9}}
th.riskh{{background:#fff2cc}}
tr:hover{{background:#f8fbff}}
.pill{{display:inline-block;padding:4px 8px;border-radius:999px;font-size:12px;font-weight:700;line-height:1.2}}
.high{{background:#dcfce7;color:#166534}}
.mid{{background:#fef3c7;color:#92400e}}
.risk{{background:#fee2e2;color:#991b1b}}
.neutral{{background:#e5e7eb;color:#334155}}
.mono{{font-family:Consolas,monospace}}
.stack{{display:grid;gap:6px}}
.line{{line-height:1.35}}
.line strong{{color:#0f172a}}
.tags{{display:flex;flex-wrap:wrap;gap:6px}}
.tag{{display:inline-block;padding:4px 8px;border-radius:10px;font-size:11px;font-weight:700}}
.tag.added{{background:#dbeafe;color:#1d4ed8}}
.tag.changed{{background:#fde68a;color:#92400e}}
.tag.warn{{background:#fee2e2;color:#991b1b}}
.tag.neutral{{background:#e5e7eb;color:#334155}}
.cell-block{{min-width:220px}}
.small{{font-size:12px}}
@media (max-width: 1200px) {{
  .filters,.cards{{grid-template-columns:repeat(2,minmax(180px,1fr))}}
}}
@media (max-width: 700px) {{
  .filters,.cards{{grid-template-columns:1fr}}
  .main{{padding:14px}}
  .page-title{{font-size:26px}}
}}
</style>
</head>
<body>
<div class="app">
  <main class="main">
    <h1 class="page-title">Painel de Fornecedores</h1>
    <div class="filters">
      <input id="q" placeholder="Buscar por codigo, CNPJ, nome">
      <select id="empresa"></select>
      <select id="movimento"></select>
      <select id="credito"></select>
    </div>
    <div class="cards">
      <div class="card"><div class="label">Linhas</div><div class="value" id="countRows">0</div></div>
      <div class="card"><div class="label">Com Movimento</div><div class="value" id="countMov">0</div></div>
      <div class="card"><div class="label">Alto Potencial</div><div class="value" id="countHigh">0</div></div>
      <div class="card"><div class="label">Validacao Fiscal</div><div class="value" id="countVal">0</div></div>
    </div>
    <div class="right-note">Cada linha mostra o cadastro original, o que foi acrescentado e as acoes sugeridas para aquele fornecedor.</div>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th class="meta">Fornecedor</th>
            <th class="before">Cadastro Original</th>
            <th class="after">Dados Incluidos</th>
            <th class="analysis">Analise Incluida</th>
            <th class="riskh">Pendencias e Acoes</th>
            <th class="riskh">O Que Mudou</th>
          </tr>
        </thead>
        <tbody id="tbody"></tbody>
      </table>
    </div>
  </main>
</div>
<script>
const DATA = {payload};
const els = {{
  q: document.getElementById('q'),
  empresa: document.getElementById('empresa'),
  movimento: document.getElementById('movimento'),
  credito: document.getElementById('credito'),
  tbody: document.getElementById('tbody'),
  countRows: document.getElementById('countRows'),
  countMov: document.getElementById('countMov'),
  countHigh: document.getElementById('countHigh'),
  countVal: document.getElementById('countVal'),
}};

function unique(values) {{
  return [...new Set(values.filter(Boolean))].sort();
}}

function fillSelect(el, values, label) {{
  el.innerHTML = `<option value="">${{label}}</option>` + values.map(v => `<option value="${{v}}">${{v}}</option>`).join('');
}}

function badgeClass(value) {{
  if (value === 'alto_potencial') return 'high';
  if (value === 'potencial_condicionado') return 'mid';
  if (value === 'baixo_ou_inexistente') return 'risk';
  return 'neutral';
}}

function ynClass(value) {{
  return value === 'S' ? 'risk' : 'high';
}}

function safe(value) {{
  return value ? value : '—';
}}

function chipsFromText(value, kind) {{
  if (!value) return '<span class="tag neutral">sem informacao</span>';
  return value.split(';').filter(Boolean).map(item => `<span class="tag ${{kind}}">${{item}}</span>`).join('');
}}

function buildChangeTags(row) {{
  const tags = [];
  if (row.depois_razao_social_oficial) tags.push('<span class="tag added">razao oficial</span>');
  if (row.depois_nome_fantasia_oficial) tags.push('<span class="tag added">fantasia oficial</span>');
  if (row.depois_email_oficial) tags.push('<span class="tag added">email oficial</span>');
  if (row.depois_status_oficial) tags.push('<span class="tag added">status oficial</span>');
  if (row.regime_indiciado) tags.push('<span class="tag added">regime</span>');
  if (row.potencial_credito_2027) tags.push('<span class="tag added">credito 2027</span>');
  if (row.matriz_filial) tags.push('<span class="tag added">matriz/filial</span>');
  if (row.porte_empresa) tags.push('<span class="tag added">porte</span>');
  if (row.natureza_juridica) tags.push('<span class="tag added">natureza juridica</span>');
  if (row.divergencia_razao_social === 'S') tags.push('<span class="tag changed">razao divergente</span>');
  if (row.divergencia_nome_fantasia === 'S') tags.push('<span class="tag changed">fantasia divergente</span>');
  if (row.divergencia_email === 'S') tags.push('<span class="tag changed">email divergente</span>');
  if (!tags.length) tags.push('<span class="tag neutral">sem inclusao capturada nesta fase</span>');
  return `<div class="tags">${{tags.join('')}}</div>`;
}}

function buildSupplierCell(row) {{
  return `
    <div class="cell-block stack">
      <div class="line"><strong>Empresa:</strong> ${{safe(row.empresa_origem)}}</div>
      <div class="line"><strong>Codigo:</strong> <span class="mono">${{safe(row.codigo_fornecedor)}}</span></div>
      <div class="line"><strong>Documento:</strong> <span class="mono">${{safe(row.documento)}}</span></div>
      <div class="line"><strong>Tipo:</strong> ${{safe(row.tipo_cadastro)}}</div>
    </div>`;
}}

function buildBeforeCell(row) {{
  return `
    <div class="cell-block stack">
      <div class="line"><strong>Razao:</strong> ${{safe(row.antes_razao_social)}}</div>
      <div class="line"><strong>Fantasia:</strong> ${{safe(row.antes_nome_fantasia)}}</div>
      <div class="line"><strong>Status:</strong> ${{safe(row.antes_status_interno)}}</div>
      <div class="line"><strong>Emails:</strong> ${{safe(row.antes_emails)}}</div>
    </div>`;
}}

function buildAfterCell(row) {{
  return `
    <div class="cell-block stack">
      <div class="line"><strong>Razao oficial:</strong> ${{safe(row.depois_razao_social_oficial)}}</div>
      <div class="line"><strong>Fantasia oficial:</strong> ${{safe(row.depois_nome_fantasia_oficial)}}</div>
      <div class="line"><strong>Status oficial:</strong> ${{safe(row.depois_status_oficial)}}</div>
      <div class="line"><strong>Email oficial:</strong> ${{safe(row.depois_email_oficial)}}</div>
      <div class="line"><strong>Matriz/filial:</strong> ${{safe(row.matriz_filial)}}</div>
      <div class="line"><strong>Porte:</strong> ${{safe(row.porte_empresa)}}</div>
      <div class="line"><strong>Natureza:</strong> ${{safe(row.natureza_juridica)}}</div>
    </div>`;
}}

function buildAnalysisCell(row) {{
  return `
    <div class="cell-block stack">
      <div class="line"><strong>Regime:</strong> <span class="pill ${{badgeClass(row.potencial_credito_2027)}}">${{safe(row.regime_indiciado)}}</span></div>
      <div class="line"><strong>Credito 2027:</strong> <span class="pill ${{badgeClass(row.potencial_credito_2027)}}">${{safe(row.potencial_credito_2027)}}</span></div>
      <div class="line"><strong>Movimento NFE:</strong> <span class="pill ${{row.movimento_nfe === 'S' ? 'high' : 'neutral'}}">${{safe(row.movimento_nfe)}}</span></div>
      <div class="line"><strong>Linhas NFE:</strong> ${{safe(row.nfe_linhas)}}</div>
      <div class="line"><strong>Valor NFE:</strong> ${{safe(row.nfe_valor_total)}}</div>
      <div class="line"><strong>Curva:</strong> ${{safe(row.curva_forn)}}</div>
      <div class="line"><strong>Score cadastral:</strong> ${{safe(row.score_cadastral)}} (${{safe(row.faixa_cadastral)}})</div>
    </div>`;
}}

function buildActionCell(row) {{
  return `
    <div class="cell-block stack">
      <div class="line"><strong>Saneamento:</strong> <span class="pill ${{ynClass(row.necessita_saneamento_cadastral)}}">${{safe(row.necessita_saneamento_cadastral)}}</span></div>
      <div class="tags small">${{chipsFromText(row.motivos_saneamento_cadastral, 'warn')}}</div>
      <div class="line"><strong>Validacao fiscal:</strong> <span class="pill ${{ynClass(row.necessita_validacao_fiscal)}}">${{safe(row.necessita_validacao_fiscal)}}</span></div>
      <div class="tags small">${{chipsFromText(row.motivos_validacao_fiscal, 'warn')}}</div>
    </div>`;
}}

fillSelect(els.empresa, unique(DATA.map(r => r.empresa_origem)), 'Todas as empresas');
fillSelect(els.movimento, ['', 'S', 'N'], 'Movimento NFE');
fillSelect(els.credito, unique(DATA.map(r => r.potencial_credito_2027)), 'Todos os creditos');

function filtered() {{
  const q = els.q.value.trim().toLowerCase();
  return DATA.filter(row => {{
    if (els.empresa.value && row.empresa_origem !== els.empresa.value) return false;
    if (els.movimento.value && row.movimento_nfe !== els.movimento.value) return false;
    if (els.credito.value && row.potencial_credito_2027 !== els.credito.value) return false;
    if (!q) return true;
    const hay = [
      row.codigo_fornecedor, row.documento, row.antes_razao_social, row.depois_razao_social_oficial,
      row.antes_nome_fantasia, row.depois_nome_fantasia_oficial, row.depois_email_oficial,
      row.regime_indiciado, row.potencial_credito_2027, row.motivos_saneamento_cadastral,
      row.motivos_validacao_fiscal
    ].join(' ').toLowerCase();
    return hay.includes(q);
  }});
}}

function render() {{
  const rows = filtered();
  els.countRows.textContent = rows.length;
  els.countMov.textContent = rows.filter(r => r.movimento_nfe === 'S').length;
  els.countHigh.textContent = rows.filter(r => r.potencial_credito_2027 === 'alto_potencial').length;
  els.countVal.textContent = rows.filter(r => r.necessita_validacao_fiscal === 'S').length;
  els.tbody.innerHTML = rows.slice(0, 1000).map((row, index) => `
    <tr data-index="${{index}}">
      <td>${{buildSupplierCell(row)}}</td>
      <td>${{buildBeforeCell(row)}}</td>
      <td>${{buildAfterCell(row)}}</td>
      <td>${{buildAnalysisCell(row)}}</td>
      <td>${{buildActionCell(row)}}</td>
      <td>${{buildChangeTags(row)}}</td>
    </tr>`).join('');
}}

[els.q, els.empresa, els.movimento, els.credito].forEach(el => el.addEventListener('input', render));
render();
</script>
</body>
</html>"""
    HTML_FILE.write_text(html, encoding="utf-8")


def build_visual_rows() -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    import csv

    normalized_map = load_csv_index(
        NORMALIZED_FILE,
        ["empresa_origem", "codigo_fornecedor", "cnpj", "razao_social"],
    )
    reconciled_by_key = load_csv_by_field(RECONCILED_FILE, "chave_unificacao")
    classified_by_key = load_csv_by_field(CLASSIFIED_FILE, "chave_unificacao")

    workbook = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    rows_by_sheet: dict[str, list[dict[str, Any]]] = {}
    all_rows: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        iterator = ws.iter_rows(values_only=True)
        header = next(iterator)
        idx = {name: pos for pos, name in enumerate(header)}
        sheet_rows: list[dict[str, Any]] = []
        for excel_row_idx, row in enumerate(iterator, start=2):
            codigo = normalize_text_upper(row[idx["CDFORNECED"]])
            cnpj = normalize_cnpj(row[idx["NRINSJURFORN"]])
            razao = normalize_text_upper(row[idx["NMRAZSOCFORN"]])
            key = build_key(sheet_name, codigo, cnpj, razao)
            normalized = normalized_map.get(key, {})
            chave_unificacao = normalized.get("chave_unificacao", "")
            reconciled = reconciled_by_key.get(chave_unificacao, {})
            classified = classified_by_key.get(chave_unificacao, {})

            visual = {
                "linha_original": excel_row_idx,
                "empresa_origem": sheet_name,
                "codigo_fornecedor": codigo,
                "tipo_cadastro": classify_code_type(codigo),
                "documento": cnpj or (codigo[1:] if len(codigo) > 1 else ""),
                "chave_unificacao": chave_unificacao,
                "antes_razao_social": razao,
                "depois_razao_social_oficial": classified.get("razao_social_oficial", ""),
                "antes_nome_fantasia": normalize_text_upper(row[idx["NMFANTFORN"]]),
                "depois_nome_fantasia_oficial": "",
                "antes_status_interno": normalize_text_upper(row[idx["IDSITUCADA"]]),
                "depois_status_oficial": classified.get("situacao_cadastral_oficial", ""),
                "antes_emails": normalized.get("emails", ""),
                "depois_email_oficial": "",
                "movimento_nfe": reconciled.get("tem_movimento_nfe", ""),
                "nfe_linhas": reconciled.get("nfe_linhas", ""),
                "nfe_valor_total": reconciled.get("nfe_valor_total", ""),
                "curva_forn": reconciled.get("curva_classificacao", ""),
                "score_cadastral": reconciled.get("score_completude", normalized.get("score_completude", "")),
                "faixa_cadastral": reconciled.get("faixa_completude", normalized.get("faixa_completude", "")),
                "pendencias_essenciais": normalized.get("pendencias_essenciais", ""),
                "pendencias_complementares": normalized.get("pendencias_complementares", ""),
                "regime_indiciado": classified.get("regime_indiciado", ""),
                "potencial_credito_2027": classified.get("potencial_credito_2027_inicial", ""),
                "necessita_saneamento_cadastral": classified.get("necessita_saneamento_cadastral", ""),
                "motivos_saneamento_cadastral": classified.get("motivos_saneamento_cadastral", ""),
                "necessita_validacao_fiscal": classified.get("necessita_validacao_fiscal", ""),
                "motivos_validacao_fiscal": classified.get("motivos_validacao_fiscal", ""),
                "divergencia_razao_social": classified.get("divergencia_razao_social_oficial", reconciled.get("divergencia_razao_social", "")),
                "divergencia_nome_fantasia": classified.get("divergencia_nome_fantasia_oficial", reconciled.get("divergencia_nome_fantasia", "")),
                "divergencia_email": classified.get("divergencia_email_oficial", ""),
                "matriz_filial": classified.get("matriz_filial", ""),
                "porte_empresa": classified.get("porte_empresa", ""),
                "natureza_juridica": classified.get("natureza_juridica", ""),
                "prioridade_saneamento_score": reconciled.get("prioridade_saneamento_score", ""),
                "prioridade_saneamento_faixa": reconciled.get("prioridade_saneamento_faixa", ""),
            }
            # enrich from fase 3 raw file for official fantasy/email if available
            # fallback via classified rows is limited, so leave blanks when absent
            sheet_rows.append(visual)
            all_rows.append(visual)
        rows_by_sheet[sheet_name] = sheet_rows

    # load official details from phase 3 enriched raw file to complete fantasy/email
    enriched_raw = load_csv_by_field(ENRICHED_RAW_FILE, "chave_unificacao")
    for row in all_rows:
        raw = enriched_raw.get(row["chave_unificacao"], {})
        row["depois_nome_fantasia_oficial"] = raw.get("nome_fantasia_oficial", "")
        row["depois_email_oficial"] = raw.get("email_oficial", "")
    return rows_by_sheet, all_rows


def main() -> None:
    PHASE_04_DIR.mkdir(parents=True, exist_ok=True)
    rows_by_sheet, all_rows = build_visual_rows()
    build_compare_workbook(rows_by_sheet)
    build_return_workbook(rows_by_sheet)
    build_html_v2(all_rows)
    print(f"Arquivos gerados em: {PHASE_04_DIR}")
    print("- 01_antes_depois_visual_fornecedores.xlsx")
    print("- 02_fornecedores_formato_original_com_campos_gvg.xlsx")
    print("- 03_painel_interativo_fornecedores.html")


if __name__ == "__main__":
    main()
