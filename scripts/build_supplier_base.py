from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
PHASE_OUTPUT_DIR = BASE_DIR / "output" / "fase_01_cadastro"
MASTER_FILE = DATA_DIR / "FORNECEDORES - TODAS AS EMPRESAS.xlsx"

EMAIL_FIELDS = [
    "DSEMAILCOT",
    "DSMAILCOTCOP",
    "DSEMAILPEDCOLFO",
    "DSMAILPEDCOLCOP",
    "DSEMAILPGTOFOR",
    "DSEMAILAFFOR",
    "DSEMAILAFFO",
    "DSEMAILAFCOP",
    "DSEMAILCRFO",
    "DSEMAILCRCOP",
    "DSEMAILNOTIFATENDAF",
]

RECORD_COLUMNS = [
    "chave_unificacao",
    "empresa_origem",
    "codigo_fornecedor",
    "cnpj",
    "razao_social",
    "nome_fantasia",
    "inscricao_estadual",
    "inscricao_municipal",
    "situacao_cadastral_interna",
    "data_cadastro",
    "data_ultima_atualizacao",
    "data_ultima_compra",
    "emails",
    "quantidade_emails",
    "telefone_suporte",
    "site",
    "grupo_fiscal",
    "indicador_contribuinte_icms",
    "forma_pagamento",
    "tipo_fornecedor",
    "score_completude",
    "faixa_completude",
    "possui_pendencia_essencial",
    "pendencias_essenciais",
    "pendencias_complementares",
    "pendencias",
]

UNIFIED_COLUMNS = [
    "chave_unificacao",
    "cnpj",
    "razao_social",
    "nome_fantasia",
    "empresas_origem",
    "quantidade_empresas_origem",
    "codigos_fornecedor",
    "quantidade_codigos_fornecedor",
    "situacao_cadastral_interna",
    "inscricao_estadual",
    "inscricao_municipal",
    "data_cadastro",
    "data_ultima_atualizacao",
    "data_ultima_compra",
    "emails",
    "quantidade_emails",
    "telefone_suporte",
    "site",
    "grupo_fiscal",
    "indicador_contribuinte_icms",
    "forma_pagamento",
    "tipo_fornecedor",
    "score_completude",
    "faixa_completude",
    "possui_pendencia_essencial",
    "pendencias_essenciais",
    "pendencias_complementares",
    "pendencias",
    "divergencia_razao_social",
    "divergencia_nome_fantasia",
    "divergencia_status_interno",
    "divergencia_ie",
    "divergencia_im",
    "registro_referencia_empresa",
    "registro_referencia_codigo",
]

PENDENCY_COLUMNS = [
    "chave_unificacao",
    "cnpj",
    "razao_social",
    "faixa_completude",
    "score_completude",
    "possui_pendencia_essencial",
    "pendencias_essenciais",
    "pendencias_complementares",
    "pendencias",
    "empresas_origem",
    "codigos_fornecedor",
    "situacao_cadastral_interna",
    "emails",
    "telefone_suporte",
    "site",
    "grupo_fiscal",
    "indicador_contribuinte_icms",
    "divergencia_razao_social",
    "divergencia_nome_fantasia",
    "divergencia_status_interno",
    "divergencia_ie",
    "divergencia_im",
]

SCORE_WEIGHTS = {
    "cnpj": 20,
    "razao_social": 15,
    "nome_fantasia": 5,
    "emails": 15,
    "inscricao_estadual": 10,
    "inscricao_municipal": 5,
    "telefone_suporte": 5,
    "site": 5,
    "situacao_cadastral_interna": 5,
    "data_cadastro": 5,
    "grupo_fiscal": 5,
    "indicador_contribuinte_icms": 5,
}

ESSENTIAL_PENDENCIES = {
    "cnpj_invalido",
    "razao_social",
    "email",
    "inscricao_estadual",
    "inscricao_municipal",
    "situacao_interna_nao_ativa",
    "data_cadastro",
}


def normalize_cnpj(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits if len(digits) == 14 else ""


def normalize_text(value: Any) -> str:
    text = str(value or "").strip()
    return re.sub(r"\s+", " ", text)


def normalize_text_upper(value: Any) -> str:
    return normalize_text(value).upper()


def to_iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value)
    return text


def extract_emails(*values: Any) -> list[str]:
    found: list[str] = []
    seen: set[str] = set()
    pattern = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)
    for value in values:
        for email in pattern.findall(str(value or "")):
            normalized = email.strip().lower()
            if normalized and normalized not in seen:
                seen.add(normalized)
                found.append(normalized)
    return found


def pick_first(records: list[dict[str, Any]], field: str) -> str:
    for record in records:
        value = normalize_text(record.get(field))
        if value:
            return value
    return ""


def summarize_status(records: list[dict[str, Any]]) -> str:
    statuses = sorted({normalize_text(record.get("situacao_cadastral_interna")) for record in records if normalize_text(record.get("situacao_cadastral_interna"))})
    if not statuses:
        return ""
    if len(statuses) == 1:
        return statuses[0]
    return "MISTO"


def yes_no(value: bool) -> str:
    return "S" if value else "N"


def classify_score(score: int) -> str:
    if score >= 85:
        return "alta"
    if score >= 60:
        return "media"
    return "baixa"


def build_unification_key(company: str, supplier_code: str, cnpj: str) -> str:
    if cnpj:
        return cnpj
    return f"SEM_CNPJ::{company}::{supplier_code}"


def calculate_score(payload: dict[str, Any]) -> tuple[int, list[str]]:
    score = 0
    missing: list[str] = []

    if normalize_cnpj(payload.get("cnpj")):
        score += SCORE_WEIGHTS["cnpj"]
    else:
        missing.append("cnpj_invalido")

    if normalize_text(payload.get("razao_social")):
        score += SCORE_WEIGHTS["razao_social"]
    else:
        missing.append("razao_social")

    if normalize_text(payload.get("nome_fantasia")):
        score += SCORE_WEIGHTS["nome_fantasia"]
    else:
        missing.append("nome_fantasia")

    if normalize_text(payload.get("emails")):
        score += SCORE_WEIGHTS["emails"]
    else:
        missing.append("email")

    if normalize_text(payload.get("inscricao_estadual")):
        score += SCORE_WEIGHTS["inscricao_estadual"]
    else:
        missing.append("inscricao_estadual")

    if normalize_text(payload.get("inscricao_municipal")):
        score += SCORE_WEIGHTS["inscricao_municipal"]
    else:
        missing.append("inscricao_municipal")

    if normalize_text(payload.get("telefone_suporte")):
        score += SCORE_WEIGHTS["telefone_suporte"]
    else:
        missing.append("telefone_suporte")

    if normalize_text(payload.get("site")):
        score += SCORE_WEIGHTS["site"]
    else:
        missing.append("site")

    if normalize_text(payload.get("situacao_cadastral_interna")) == "A":
        score += SCORE_WEIGHTS["situacao_cadastral_interna"]
    else:
        missing.append("situacao_interna_nao_ativa")

    if normalize_text(payload.get("data_cadastro")):
        score += SCORE_WEIGHTS["data_cadastro"]
    else:
        missing.append("data_cadastro")

    if normalize_text(payload.get("grupo_fiscal")):
        score += SCORE_WEIGHTS["grupo_fiscal"]
    else:
        missing.append("grupo_fiscal")

    if normalize_text(payload.get("indicador_contribuinte_icms")):
        score += SCORE_WEIGHTS["indicador_contribuinte_icms"]
    else:
        missing.append("indicador_contribuinte_icms")

    return score, missing


def split_missing_fields(missing: list[str]) -> tuple[list[str], list[str]]:
    essential = [field for field in missing if field in ESSENTIAL_PENDENCIES]
    complementary = [field for field in missing if field not in ESSENTIAL_PENDENCIES]
    return essential, complementary


def load_records() -> list[dict[str, Any]]:
    workbook = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        index = {name: position for position, name in enumerate(header)}

        for row in rows:
            email_values = [row[index[field]] for field in EMAIL_FIELDS if field in index]
            emails = extract_emails(*email_values)

            record = {
                "chave_unificacao": build_unification_key(
                    sheet_name,
                    normalize_text_upper(row[index["CDFORNECED"]]),
                    normalize_cnpj(row[index["NRINSJURFORN"]]),
                ),
                "empresa_origem": sheet_name,
                "codigo_fornecedor": normalize_text_upper(row[index["CDFORNECED"]]),
                "cnpj": normalize_cnpj(row[index["NRINSJURFORN"]]),
                "razao_social": normalize_text_upper(row[index["NMRAZSOCFORN"]]),
                "nome_fantasia": normalize_text_upper(row[index["NMFANTFORN"]]),
                "inscricao_estadual": normalize_text_upper(row[index["NRINSESTFORN"]]),
                "inscricao_municipal": normalize_text_upper(row[index["NRINSMUNFORN"]]),
                "situacao_cadastral_interna": normalize_text_upper(row[index["IDSITUCADA"]]),
                "data_cadastro": to_iso_date(row[index["DTCADAFORN"]]),
                "data_ultima_atualizacao": to_iso_date(row[index["DTULTATUFORN"]]),
                "data_ultima_compra": to_iso_date(row[index["DTULTCMPFORN"]]),
                "emails": ";".join(emails),
                "quantidade_emails": len(emails),
                "telefone_suporte": normalize_text(row[index["NRTELSUPORTEFOR"]]),
                "site": normalize_text(row[index["DSHOMEPAFOR"]]).lower(),
                "grupo_fiscal": normalize_text_upper(row[index["CDGRUPOFISC"]]),
                "indicador_contribuinte_icms": normalize_text_upper(row[index["IDCONTRIBICMS"]]),
                "forma_pagamento": normalize_text_upper(row[index["CDFORMPGTO"]]),
                "tipo_fornecedor": normalize_text_upper(row[index["IDTIPOFORN"]]),
            }

            score, missing = calculate_score(record)
            essential_missing, complementary_missing = split_missing_fields(missing)
            record["score_completude"] = score
            record["faixa_completude"] = classify_score(score)
            record["possui_pendencia_essencial"] = yes_no(bool(essential_missing))
            record["pendencias_essenciais"] = ";".join(essential_missing)
            record["pendencias_complementares"] = ";".join(complementary_missing)
            record["pendencias"] = ";".join(missing)
            records.append(record)

    return records


def build_unified_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        grouped[record["chave_unificacao"]].append(record)

    unified_rows: list[dict[str, Any]] = []

    for key, group in grouped.items():
        ordered = sorted(
            group,
            key=lambda item: (
                int(item["score_completude"]),
                item["situacao_cadastral_interna"] == "A",
                item["data_ultima_atualizacao"],
                item["data_cadastro"],
                item["empresa_origem"],
            ),
            reverse=True,
        )

        best = ordered[0]
        email_set: list[str] = []
        seen_emails: set[str] = set()
        for record in ordered:
            for email in str(record.get("emails") or "").split(";"):
                normalized = email.strip().lower()
                if normalized and normalized not in seen_emails:
                    seen_emails.add(normalized)
                    email_set.append(normalized)

        payload = {
            "chave_unificacao": key,
            "cnpj": best["cnpj"],
            "razao_social": pick_first(ordered, "razao_social"),
            "nome_fantasia": pick_first(ordered, "nome_fantasia"),
            "empresas_origem": ";".join(sorted({record["empresa_origem"] for record in ordered})),
            "quantidade_empresas_origem": len({record["empresa_origem"] for record in ordered}),
            "codigos_fornecedor": ";".join(sorted({record["codigo_fornecedor"] for record in ordered if record["codigo_fornecedor"]})),
            "quantidade_codigos_fornecedor": len({record["codigo_fornecedor"] for record in ordered if record["codigo_fornecedor"]}),
            "situacao_cadastral_interna": summarize_status(ordered),
            "inscricao_estadual": pick_first(ordered, "inscricao_estadual"),
            "inscricao_municipal": pick_first(ordered, "inscricao_municipal"),
            "data_cadastro": pick_first(ordered, "data_cadastro"),
            "data_ultima_atualizacao": pick_first(ordered, "data_ultima_atualizacao"),
            "data_ultima_compra": pick_first(ordered, "data_ultima_compra"),
            "emails": ";".join(email_set),
            "quantidade_emails": len(email_set),
            "telefone_suporte": pick_first(ordered, "telefone_suporte"),
            "site": pick_first(ordered, "site"),
            "grupo_fiscal": pick_first(ordered, "grupo_fiscal"),
            "indicador_contribuinte_icms": pick_first(ordered, "indicador_contribuinte_icms"),
            "forma_pagamento": pick_first(ordered, "forma_pagamento"),
            "tipo_fornecedor": pick_first(ordered, "tipo_fornecedor"),
            "divergencia_razao_social": yes_no(len({record["razao_social"] for record in ordered if record["razao_social"]}) > 1),
            "divergencia_nome_fantasia": yes_no(len({record["nome_fantasia"] for record in ordered if record["nome_fantasia"]}) > 1),
            "divergencia_status_interno": yes_no(len({record["situacao_cadastral_interna"] for record in ordered if record["situacao_cadastral_interna"]}) > 1),
            "divergencia_ie": yes_no(len({record["inscricao_estadual"] for record in ordered if record["inscricao_estadual"]}) > 1),
            "divergencia_im": yes_no(len({record["inscricao_municipal"] for record in ordered if record["inscricao_municipal"]}) > 1),
            "registro_referencia_empresa": best["empresa_origem"],
            "registro_referencia_codigo": best["codigo_fornecedor"],
        }

        score, missing = calculate_score(payload)
        essential_missing, complementary_missing = split_missing_fields(missing)
        payload["score_completude"] = score
        payload["faixa_completude"] = classify_score(score)
        payload["possui_pendencia_essencial"] = yes_no(bool(essential_missing))
        payload["pendencias_essenciais"] = ";".join(essential_missing)
        payload["pendencias_complementares"] = ";".join(complementary_missing)
        payload["pendencias"] = ";".join(missing)
        unified_rows.append(payload)

    return sorted(unified_rows, key=lambda item: (item["score_completude"], item["razao_social"], item["cnpj"]))


def build_pendency_rows(unified_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in unified_rows:
        has_divergence = any(row.get(field) == "S" for field in [
            "divergencia_razao_social",
            "divergencia_nome_fantasia",
            "divergencia_status_interno",
            "divergencia_ie",
            "divergencia_im",
        ])
        if row["pendencias"] or has_divergence:
            rows.append({key: row.get(key, "") for key in PENDENCY_COLUMNS})
    return sorted(rows, key=lambda item: (int(item["score_completude"]), item["razao_social"], item["cnpj"]))


def write_csv(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def build_summary(records: list[dict[str, Any]], unified_rows: list[dict[str, Any]], pendencies: list[dict[str, Any]]) -> dict[str, Any]:
    missing_counter: Counter[str] = Counter()
    essential_missing_counter: Counter[str] = Counter()
    complementary_missing_counter: Counter[str] = Counter()
    faixa_counter: Counter[str] = Counter()
    source_counter: Counter[str] = Counter()
    divergence_counter: Counter[str] = Counter()

    for row in unified_rows:
        faixa_counter[str(row["faixa_completude"])] += 1
        for company in str(row["empresas_origem"] or "").split(";"):
            if company:
                source_counter[company] += 1
        for field in str(row["pendencias"] or "").split(";"):
            if field:
                missing_counter[field] += 1
        for field in str(row["pendencias_essenciais"] or "").split(";"):
            if field:
                essential_missing_counter[field] += 1
        for field in str(row["pendencias_complementares"] or "").split(";"):
            if field:
                complementary_missing_counter[field] += 1
        for field in [
            "divergencia_razao_social",
            "divergencia_nome_fantasia",
            "divergencia_status_interno",
            "divergencia_ie",
            "divergencia_im",
        ]:
            if row.get(field) == "S":
                divergence_counter[field] += 1

    return {
        "fontes": {
            "cadastro_mestre": str(MASTER_FILE),
        },
        "gerado_em": datetime.now().isoformat(timespec="seconds"),
        "totais": {
            "registros_normalizados": len(records),
            "fornecedores_unificados": len(unified_rows),
            "fornecedores_com_pendencias_ou_divergencias": len(pendencies),
            "fornecedores_com_cnpj_valido": sum(1 for row in unified_rows if row["cnpj"]),
            "fornecedores_sem_cnpj_valido": sum(1 for row in unified_rows if not row["cnpj"]),
            "fornecedores_com_pendencia_essencial": sum(1 for row in unified_rows if row["possui_pendencia_essencial"] == "S"),
        },
        "faixas_completude": dict(faixa_counter),
        "pendencias_mais_frequentes": dict(missing_counter.most_common()),
        "pendencias_essenciais_mais_frequentes": dict(essential_missing_counter.most_common()),
        "pendencias_complementares_mais_frequentes": dict(complementary_missing_counter.most_common()),
        "cobertura_por_empresa_origem": dict(source_counter),
        "divergencias": dict(divergence_counter),
        "score_medio_unificado": round(sum(int(row["score_completude"]) for row in unified_rows) / max(1, len(unified_rows)), 2),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera a base cadastral normalizada e unificada de fornecedores.")
    parser.add_argument("--output-dir", default=str(PHASE_OUTPUT_DIR), help="Diretorio de saida para os artefatos.")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    records = load_records()
    unified_rows = build_unified_records(records)
    pendency_rows = build_pendency_rows(unified_rows)
    summary = build_summary(records, unified_rows, pendency_rows)

    write_csv(output_dir / "01_fornecedores_cadastro_normalizado.csv", RECORD_COLUMNS, records)
    write_csv(output_dir / "02_fornecedores_cadastro_unificado.csv", UNIFIED_COLUMNS, unified_rows)
    write_csv(output_dir / "03_fornecedores_pendencias_cadastrais.csv", PENDENCY_COLUMNS, pendency_rows)

    with (output_dir / "04_fornecedores_resumo_cadastral.json").open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, ensure_ascii=False, indent=2)

    print(f"Arquivos gerados em: {output_dir}")
    print(f"- 01_fornecedores_cadastro_normalizado.csv: {len(records)} linhas")
    print(f"- 02_fornecedores_cadastro_unificado.csv: {len(unified_rows)} linhas")
    print(f"- 03_fornecedores_pendencias_cadastrais.csv: {len(pendency_rows)} linhas")
    print(f"- 04_fornecedores_resumo_cadastral.json")


if __name__ == "__main__":
    main()
