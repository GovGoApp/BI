from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
MASTER_FILE = DATA_DIR / "FORNECEDORES - TODAS AS EMPRESAS.xlsx"
CURVE_FILE = DATA_DIR / "CURVA_FORN_-_TODAS.csv"
NFE_FILE = DATA_DIR / "NFE.csv"


def normalize_cnpj(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits if len(digits) == 14 else ""


def pct(part: int, total: int) -> float:
    if total <= 0:
        return 0.0
    return round((part / total) * 100, 2)


def inspect_master() -> dict[str, Any]:
    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    per_sheet: dict[str, Any] = {}
    all_codes: set[str] = set()
    all_cnpjs: set[str] = set()
    cnpj_companies: dict[str, set[str]] = {}
    totals = {
        "rows": 0,
        "missing_email": 0,
        "missing_ie": 0,
        "missing_im": 0,
    }
    status_counter: Counter[str] = Counter()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        idx = {name: position for position, name in enumerate(header)}
        sheet_total = 0
        missing_email = 0
        missing_ie = 0
        missing_im = 0

        for row in rows:
            sheet_total += 1
            totals["rows"] += 1

            code = str(row[idx["CDFORNECED"]] or "").strip()
            cnpj = normalize_cnpj(row[idx["NRINSJURFORN"]])
            status = str(row[idx["IDSITUCADA"]] or "").strip() or "<vazio>"
            email = row[idx["DSEMAILCOT"]]
            ie = row[idx["NRINSESTFORN"]]
            im = row[idx["NRINSMUNFORN"]]

            if code:
                all_codes.add(code)
            if cnpj:
                all_cnpjs.add(cnpj)
                cnpj_companies.setdefault(cnpj, set()).add(sheet_name)

            status_counter[status] += 1

            if not email:
                missing_email += 1
                totals["missing_email"] += 1
            if not ie:
                missing_ie += 1
                totals["missing_ie"] += 1
            if not im:
                missing_im += 1
                totals["missing_im"] += 1

        per_sheet[sheet_name] = {
            "rows": sheet_total,
            "missing_email": missing_email,
            "missing_ie": missing_ie,
            "missing_im": missing_im,
            "missing_email_pct": pct(missing_email, sheet_total),
            "missing_ie_pct": pct(missing_ie, sheet_total),
            "missing_im_pct": pct(missing_im, sheet_total),
        }

    company_distribution = Counter(len(companies) for companies in cnpj_companies.values())

    return {
        "sheets": list(wb.sheetnames),
        "per_sheet": per_sheet,
        "unique_codes": len(all_codes),
        "unique_cnpjs": len(all_cnpjs),
        "shared_cnpjs": sum(1 for companies in cnpj_companies.values() if len(companies) > 1),
        "company_distribution_by_cnpj": dict(sorted(company_distribution.items())),
        "status": dict(status_counter),
        "totals": {
            **totals,
            "missing_email_pct": pct(totals["missing_email"], totals["rows"]),
            "missing_ie_pct": pct(totals["missing_ie"], totals["rows"]),
            "missing_im_pct": pct(totals["missing_im"], totals["rows"]),
        },
    }


def inspect_curve() -> dict[str, Any]:
    curves: Counter[str] = Counter()
    codes: set[str] = set()

    with CURVE_FILE.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            code = str(row["CDFORNECED"] or "").strip()
            curve = str(row["CURVA"] or "").strip() or "<vazio>"
            if code:
                codes.add(code)
            curves[curve] += 1

    return {
        "rows": sum(curves.values()),
        "unique_codes": len(codes),
        "curves": dict(curves),
        "codes": codes,
    }


def inspect_nfe() -> dict[str, Any]:
    company_counter: Counter[str] = Counter()
    uf_counter: Counter[str] = Counter()
    curve_counter: Counter[str] = Counter()
    supplier_codes: set[str] = set()
    official_supplier_codes: set[str] = set()
    products: set[str] = set()
    rows = 0

    with NFE_FILE.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            rows += 1
            company_counter[str(row["NMEMP"] or "").strip() or "<vazio>"] += 1
            uf_counter[str(row["UF"] or "").strip() or "<vazio>"] += 1
            curve_counter[str(row["CURVA_FORN"] or "").strip() or "<vazio>"] += 1

            code = str(row["CDFORNECED"] or "").strip()
            official_code = str(row["CDFORNECED_OFICIAL"] or "").strip()
            product = str(row["CDPRODUTO_OFICIAL"] or row["CDPRODUTO"] or "").strip()

            if code:
                supplier_codes.add(code)
            if official_code:
                official_supplier_codes.add(official_code)
            if product:
                products.add(product)

    return {
        "rows": rows,
        "companies": dict(company_counter),
        "ufs": dict(uf_counter),
        "curve_forn": dict(curve_counter),
        "unique_supplier_codes": len(supplier_codes),
        "unique_official_supplier_codes": len(official_supplier_codes),
        "unique_products": len(products),
        "codes": supplier_codes,
        "official_codes": official_supplier_codes,
    }


def build_report() -> dict[str, Any]:
    master = inspect_master()
    curve = inspect_curve()
    nfe = inspect_nfe()

    master_codes = set()
    for sheet in master["per_sheet"].values():
        _ = sheet
    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        idx = {name: position for position, name in enumerate(header)}
        for row in rows:
            code = str(row[idx["CDFORNECED"]] or "").strip()
            if code:
                master_codes.add(code)

    nfe_union = nfe["codes"] | nfe["official_codes"]
    curve_codes = curve["codes"]

    overlaps = {
        "curve_in_master": len(curve_codes & master_codes),
        "curve_not_in_master": len(curve_codes - master_codes),
        "nfe_in_master": len(nfe_union & master_codes),
        "nfe_not_in_master": len(nfe_union - master_codes),
        "master_without_nfe": len(master_codes - nfe_union),
        "master_without_curve": len(master_codes - curve_codes),
    }

    return {
        "master": {key: value for key, value in master.items() if key != "codes"},
        "curve": {key: value for key, value in curve.items() if key != "codes"},
        "nfe": {key: value for key, value in nfe.items() if key not in {"codes", "official_codes"}},
        "overlaps": overlaps,
    }


def render_text(report: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append("Inventario das bases")
    lines.append("")
    lines.append(f"Cadastro mestre: {report['master']['totals']['rows']} linhas")
    lines.append(f"CNPJs unicos: {report['master']['unique_cnpjs']}")
    lines.append(f"Curva: {report['curve']['rows']} linhas")
    lines.append(f"NFe: {report['nfe']['rows']} linhas")
    lines.append("")
    lines.append("Cobertura")
    lines.append(f"- fornecedores da curva no cadastro: {report['overlaps']['curve_in_master']}")
    lines.append(f"- fornecedores da curva fora do cadastro: {report['overlaps']['curve_not_in_master']}")
    lines.append(f"- fornecedores da NFe no cadastro: {report['overlaps']['nfe_in_master']}")
    lines.append(f"- fornecedores da NFe fora do cadastro: {report['overlaps']['nfe_not_in_master']}")
    lines.append(f"- fornecedores do cadastro sem NFe: {report['overlaps']['master_without_nfe']}")
    lines.append(f"- fornecedores do cadastro sem curva: {report['overlaps']['master_without_curve']}")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera inventario inicial das bases de fornecedores.")
    parser.add_argument("--format", choices=["json", "text"], default="text")
    args = parser.parse_args()

    report = build_report()

    if args.format == "json":
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return

    print(render_text(report))


if __name__ == "__main__":
    main()
